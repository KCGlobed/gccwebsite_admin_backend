from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import *
from .serializers import *
from rest_framework import filters
from gcc_backend.pagination import CustomPageNumberPagination
from rest_framework.permissions import IsAuthenticated
from django.utils.dateparse import parse_date
from gcc_backend.utils import *
from gcc_backend import settings
from google.cloud import storage
import pandas as pd
import tempfile
import re
from datetime import datetime, timedelta
from django.utils.dateparse import parse_date
client = storage.Client(project=settings.GS_PROJECT_ID)
import os

from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template

from gcc_backend.utils import send_email_async
import threading
from django.conf import settings
from django.db.models import OuterRef, Exists

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from students.models import *
from datetime import date, timedelta, datetime
from django.db.models import Count
from django.db.models.functions import TruncDate
from users.service import zoom

from .models import DossierData
from django.utils import timezone


class CareerApplication_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id']
    ordering_fields = ['id']
    def get(self, request):
        datas = CareerApplication.objects.all().order_by('-id')

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListCareerApplicationSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    

class PartnerWithUs_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id']
    ordering_fields = ['id']
    def get(self, request):
        datas = PartnerWithUs.objects.all().order_by('-id')

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListPartnerWithUsSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    

class DossierDataForm_Create(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = CreateDossierDataSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            obj = serializer.save()
            source = request.data.get("source")
            if str(source)=="16":
                pdf_url = f"{settings.STATIC_URL}files/CPA-STUDENT-DOSSIER.pdf"
            elif str(source)=="17":
                pdf_url = f"{settings.STATIC_URL}files/EA-STUDENT-DOSSIER.pdf"
            else:
                pdf_url = f"{settings.STATIC_URL}files/GCC%20SCHOOL%20Dossier.pdf"
            return success_response(message="success", data={"url":pdf_url, "id":obj.id, "data":ListDossierDataSerializer(obj).data}, status_code=status.HTTP_200_OK)
        else:
            return error_response(message="failed", data = {}, status_code=status.HTTP_400_BAD_REQUEST)


class DossierDataFormCustom_Create(APIView):
    def post(self, request, format=None):
        serializer = CreateDossierDataCustomAffliateSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            obj = serializer.save()

            start_date = datetime.now().date()
            end_date = date(2026, 8, 31)

            # Count records grouped by interview_date
            # booked_slots = (
            #     DossierData.objects
            #     .filter(interview_date__range=(start_date, end_date))
            #     .annotate(day=TruncDate("interview_date"))
            #     .values("day")
            #     .annotate(count=Count("id"))
            #     .order_by("day")
            # )
            # # Convert queryset to dictionary
            # count_map = {item["day"]: item["count"] for item in booked_slots}  

            # # Fill missing dates with 0
            # result = []
            # current = start_date

            # while current <= end_date:
            #     result.append({
            #         "date": current.strftime("%d-%m-%Y"),
            #         "count": count_map.get(current, 0)
            #     })
            #     current += timedelta(days=1)


            result = []
            booking_slots = {}
    
            current_date = start_date
            
            while current_date <= end_date:
                days_from_today = (current_date - start_date).days
            
                # Higher occupancy for nearer dates
                if days_from_today <= 5:
                    booked = random.randint(23, 28)
                elif days_from_today <= 10:
                    booked = random.randint(18, 28)
                elif days_from_today <= 20:
                    booked = random.randint(10, 22)
                else:
                    booked = random.randint(2, 15)
            
                booking_slots = {
                    # "max_slots": 30,
                    # "booked_slots": booked,
                    # "available_slots": 30 - booked,
                    # "is_full": booked == 30
                    "date": current_date.strftime("%d-%m-%Y"),
                    "count": booked
                }
            
                current_date += timedelta(days=1)
                result.append(booking_slots)
            return success_response(message="success", data={"id":obj.id, "slot_data":result}, status_code=status.HTTP_200_OK)
        else:
            return error_response(message="failed", data = [], status_code=status.HTTP_400_BAD_REQUEST)





class InterviewSlotScheduleAdmin_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["full_name","email", "phone", "interview_date"]
    ordering_fields = ["id"]
    def get(self, request):
        datas = DossierData.objects.filter(source=SourceType.Affiliate7).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)

        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)



        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListDossierDataAffliateSevenInterviewSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)






from .serializers import push_to_meritto
class DossierMeritto_CreateUpdate(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        data_list = []
        j = 0
        # yesterday = datetime.now().date() - timedelta(days=2)
        today = datetime.now().date()
        # dossier_obj = DossierData.objects.filter(created_at__date__gte='2026-05-15', source=SourceType.Website)
        # dossier_obj = dossier_obj.filter(created_at__date__lte='2026-05-18')
        dossier_obj = DossierData.objects.filter(id=request.data['lid'])
        filter_data = dossier_obj.count()
        print("count data...",filter_data)
        cc = filter_data
        # objj = list(User.objects.values_list('email', flat=True))

        # print(objj)
        # print(len(objj))
        
        
        for obj in dossier_obj:
            # if obj.email in objj:
            #     print("email.......",obj.email)
            #     push_to_meritto(obj)


            # serializer = CreateOrUpdateDossierDataMerittoSerializer(obj, data = request.data)
            # if serializer.is_valid(raise_exception = True):
            #     serializer.save()


                # data_list.append(obj.id)
                
            push_to_meritto(obj)
            data_list.append(obj.id)
            # break
            j+=1
            cc-=1
            print("++++++",j)
            print("------",cc)
        total_lead = len(data_list)

        return success_response(message="success", data={"total_lead":total_lead,"filter_data":filter_data,"ids":data_list}, status_code=status.HTTP_200_OK)




class ExcelPhoneMatchAPI(APIView):

    def post(self, request, *args, **kwargs):
        print("calling...!!")
        file = request.FILES.get('file')

        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Read Excel file
            df = pd.read_excel(file)

            # Ensure column exists (change 'phone' to your column name)
            if 'Mobile' not in df.columns:
                return Response({"error": "Column 'mobile' not found in file"}, status=400)

            # Clean phone numbers
            df['Mobile'] = df['Mobile'].astype(str).str.strip()

            phone_list = df['Mobile'].dropna().unique().tolist()
            data_list = []
            j = 0
            print("listing data.....",phone_list)
            for i in phone_list:
                print(i[4:])
                matched_data = DossierData.objects.filter(phone=i[5:], source=12).first()
                print(matched_data)
                if matched_data:
                    push_to_meritto(matched_data)
                    data_list.append(matched_data.id)
                    # break
                    print("success")
                j+=1
                print(j)
            total_lead = len(data_list)
            

            # # Query DB
            # matched_data = DossierData.objects.filter(
            #     phone__in=phone_list
            # ).values('id', 'email', 'phone', 'created_at')

            # matched_phones = {item['phone'] for item in matched_data}

            # # Logic: find unmatched numbers
            # unmatched = [p for p in phone_list if p not in matched_phones]


            # url = f"{settings.MERITO_BASE_URL}/lead/v1/createOrUpdate"

            # headers = {
            #     "Content-Type": "application/json",
            #     "secret-key": settings.MERITO_SECRETE_KEY,
            #     "access-key": settings.MERITO_ACCESS_KEY
            # }

            # payload = {
            #     "name": obj.full_name,
            #     "email": obj.email,
            #     "mobile": obj.phone,
            #     "search_criteria": "mobile",
            #     "country": "India",
            #     "source": "gccvsloptin",
            #     "cf_source": "gccvsloptin",
            #     "cf_utmsource1": str(obj.utm_source).encode("ascii", "ignore").decode().strip(),
            #     "medium": str(obj.utm_medium).encode("ascii", "ignore").decode().strip(),
            #     "campaign": str(obj.utm_campaign).encode("ascii", "ignore").decode().strip(),
            #     # "cf_payment_status": "Complete",
            # }
            # if obj.city:
            #     payload["city"] = obj.city
            # if obj.state:
            #     payload["state"] = obj.state
            # if obj.university:
            #     payload["cf_fee_waiver_category"] = obj.fee_waiver_category

            # print("merito data.......",payload)

            # try:
            #     response = requests.post(url, headers=headers, json=payload)
            #     print(response.status_code)
            #     print(response.text)
            #     return response.json()
            # except requests.exceptions.RequestException as e:
            #     print("Meritto API Error:", str(e))
            #     return None





            return Response({
                "total_uploaded": len(phone_list),
                "total_lead": total_lead,
            })

        except Exception as e:
            return Response({"error": str(e)}, status=500)







class DossierDocument_Create(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = CreateDossierDocumentSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            obj = serializer.save()
            return success_response(message="success", data={"id":obj.id}, status_code=status.HTTP_200_OK)
        else:
            return error_response(message="failed", data = {}, status_code=status.HTTP_400_BAD_REQUEST)


class DossierAbondant_Create(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = CreateDossierAbondantSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            obj = serializer.save()
            return success_response(message="success", data={"id":obj.id}, status_code=status.HTTP_200_OK)
        else:
            return error_response(message="failed", data = {}, status_code=status.HTTP_400_BAD_REQUEST)


class VslDataForm_Create(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = CreateVslDataSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            obj = serializer.save()
            return success_response(message="success", data={"id":obj.id, "data":ListDossierDataSerializer(obj).data}, status_code=status.HTTP_200_OK)
        else:
            return error_response(message="failed", data = {}, status_code=status.HTTP_400_BAD_REQUEST)


class VslFinalDataForm_Create(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = CreateVslFinalDataSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            obj = serializer.save()
            return success_response(message="success", data={"id":obj.id, "data":ListDossierDataSerializer(obj).data}, status_code=status.HTTP_200_OK)
        else:
            return error_response(message="failed", data = {}, status_code=status.HTTP_400_BAD_REQUEST)


class VslOptinDetailDataForm_Create(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = CreateVslOptinDetailDataSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            obj = serializer.save()
            return success_response(message="success", data={}, status_code=status.HTTP_200_OK)
        else:
            return error_response(message="failed", data = {}, status_code=status.HTTP_400_BAD_REQUEST)


class VslOptinDetailDataForm_Update(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = UpdateVslOptinDetailDataSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            obj = serializer.save()
            return success_response(message="success", data={}, status_code=status.HTTP_200_OK)
        else:
            return error_response(message="failed", data = {}, status_code=status.HTTP_400_BAD_REQUEST)


class DossierDataForm_List(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state","referred_code","referral_code"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]
    def get(self, request):

        document_status = request.GET.get('document_status')
        if document_status:
            datas = DossierData.objects.filter(
                    source=SourceType.Website
                ).annotate(
                    has_doc=Exists(
                        DossierDocument.objects.filter(dossier_id=OuterRef('id'))
                    )
                ).filter(has_doc=True).order_by('-id')
        else:
            datas = DossierData.objects.filter(source=SourceType.Website).order_by('-id')
        
        verify_status = request.GET.get('isVerified')
        if verify_status:
            if str(verify_status) in ["2","3"]:
                datas = datas.filter(document_status=int(verify_status))
        verify_status = request.GET.get('isVerified')
        if verify_status:
            if str(verify_status) in ["2","3"]:
                datas = datas.filter(document_status=int(verify_status))
          
        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        fee_waiver_category = request.GET.get('fee_waiver_category')
        if fee_waiver_category:
            datas = datas.filter(fee_waiver_category__icontains=fee_waiver_category)

        university = request.GET.get('university')
        if university:
            datas = datas.filter(university__icontains=university)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        program = request.GET.get('program')
        if program:
            program_list = [1,2]
            if program in program_list:
                datas = datas.filter(program=program)
            else:
                datas = datas.exclude(program__in=program_list)

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListDossierDataSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    

class AbondantDataForm_List(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone"]
    ordering_fields = ['id',"full_name","email","phone","created_at"]
    def get(self, request):
        source = request.GET.get('source')
        if source:
            datas = DossierAbondant.objects.filter(source=source).order_by('-id')
        else:
            datas = DossierAbondant.objects.all().order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListDossierAbondantSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    

class DossierDataSourceForm_List(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state","referred_code","referral_code"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]
    def get(self, request):

        source_type = request.GET.get('source')
        if source_type:
            # datas = DossierData.objects.filter(source=source_type).order_by('-id')
            document_status = request.GET.get('document_status')
            if document_status:
                datas = DossierData.objects.filter(
                        source=source_type
                    ).annotate(
                        has_doc=Exists(
                            DossierDocument.objects.filter(dossier_id=OuterRef('id'))
                        )
                    ).filter(has_doc=True).order_by('-id')
            else:
                datas = DossierData.objects.filter(source=source_type).order_by('-id')
                if str(source_type) == str(SourceType.EAWebsite):
                    access_data = settings.EAUTMSOURCE
                    if str(request.user.first_name).lower() in access_data:
                        datas = DossierData.objects.filter(source=source_type, utm_source__iexact=str(request.user.first_name).lower()).order_by('-id')
                if str(source_type) == str(SourceType.CPAWebsite):
                    access_data = settings.CPAUTMSOURCE
                    if str(request.user.first_name).lower() in access_data:
                        datas = DossierData.objects.filter(source=source_type, utm_source__iexact=str(request.user.first_name).lower()).order_by('-id')
        else:
            datas = DossierData.objects.filter(source=SourceType.Website).order_by('-id')


        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)

        if source_type == "23":
            serializers = AdminListMeetingDossierDataSerializer(page, many=True)
        else:
            serializers = ListDossierDataSerializer(page, many=True)
        print(serializers.data[0])
        return paginator.get_paginated_response(serializers.data)
    

class VSLAdvisorDataForm_List(APIView):
    # permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]
    def get(self, request):
        
        adv_datas_list = list(VslDetail.objects.filter(specialist_status=True).values_list('dossier__id',flat=True))
        # print(adv_datas_list)
        datas = DossierData.objects.filter(id__in=adv_datas_list, source=SourceType.VslOptin).order_by('-id')


        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListDossierDataSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    

class NewsletterSubscribers_List(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id']
    ordering_fields = ['id']
    def get(self, request):
        datas = NewsletterSubscribers.objects.all().order_by('-id')

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListNewsletterSubscriberSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    


class CreateSupportFormView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = CreateSupportFormSerializer(data = request.data, context={'request': request})
        if serializer.is_valid(raise_exception = True):
            obj = serializer.save()
            subject = f'Feedback - {obj.subject}'
            message = obj.message
            email_from = settings.DEFAULT_FROM_EMAIL
            recipient_list = ['info@gccschool.com','support@gccschool.com']
            # html_message = ""
            html_message = loader.render_to_string(
                    'feedback_mail.html',
                    {
                        "full_name":obj.user.first_name,
                        "email":obj.user.email,
                        "application_id":obj.user.application_id,
                        "subject":obj.subject,
                        "message":obj.message,
                        "created_at":obj.created_at
                    }
                )
            threading.Thread(
                target=send_email_async,
                args=(subject, message, email_from, recipient_list, html_message)
            ).start()
            return success_response(message="success", data={"id":obj.id, "data":CreateSupportFormSerializer(obj).data}, status_code=status.HTTP_200_OK)
        else:
            return error_response(message="failed", data = {}, status_code=status.HTTP_400_BAD_REQUEST)


class SupportForm_page_list(APIView):
    # permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id']
    ordering_fields = ['id']
    def get(self, request):
        datas = SupportForm.objects.all().order_by('-id')

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListSupportFormSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    


class GetDossierReportPDFView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]
    def get(self, request, sid=None):
        document_status = request.GET.get('document_status')
        if document_status:
            datas = DossierData.objects.filter(
                    source=SourceType.Website
                ).annotate(
                    has_doc=Exists(
                        DossierDocument.objects.filter(dossier_id=OuterRef('id'))
                    )
                ).filter(has_doc=True).order_by('-id')
        else:
            datas = DossierData.objects.filter(source=SourceType.Website).order_by('-id')
        
        verify_status = request.GET.get('isVerified')
        if verify_status:
            if str(verify_status) in ["2","3"]:
                datas = datas.filter(document_status=int(verify_status))
                
        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)



        # datas = DossierData.objects.filter(source=SourceType.Website).order_by('-id')

        # full_name = request.GET.get('full_name')
        # if full_name:
        #     datas = datas.filter(full_name__icontains=full_name)

        # email = request.GET.get('email')
        # if email:
        #     datas = datas.filter(email__icontains=email)


        # phone = request.GET.get('phone')
        # if phone:
        #     datas = datas.filter(phone__icontains=phone)


        # state = request.GET.get('state')
        # if state:
        #     datas = datas.filter(state__icontains=state)

        # city = request.GET.get('city')
        # if city:
        #     datas = datas.filter(city__icontains=city)

        # # Date range filter
        # start_date = request.GET.get('start_date')
        # end_date = request.GET.get('end_date')
        # if start_date:
        #     start_date = parse_date(start_date)
        #     if start_date:
        #         datas = datas.filter(created_at__date__gte=start_date)

        # if end_date:
        #     end_date = parse_date(end_date)
        #     if end_date:
        #         datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierDataSerializer(datas, many=True)


        data = {
                    "user_data":serializers.data,
                    "report_date": datetime.now().strftime("%d-%m-%Y, %H:%M")
                }
        

        template = get_template('pdf/dossier_report.html')
        html  = template.render(data)
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Encode HTML and create PDF
            html = html.encode('latin-1', 'replace').decode('latin-1')
            pdf = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=temp_file)

            if pdf.err:
                raise Exception("PDF generation error!")
        
        # After the 'with' block, the file is closed, but not deleted yet
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "dossier_report"
            gcs_folder_name = "media/reports/dossier/pdf"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.pdf"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted from the server's disk
            os.remove(pdf_path)
    


class GetDossierReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]

    def get(self, request, sid=None):
        
        datas = DossierData.objects.filter(source=SourceType.Website).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        program = request.GET.get('program')
        if program:
            program_list = [1,2]
            if program in program_list:
                datas = datas.filter(program=program)
            else:
                datas = datas.exclude(program__in=program_list)
        

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierDataSerializer(datas, many=True)

        lis = []
        
        lis.append({
                "name":"Dossier Report",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "fbc_id":'',
                "utm_source":'',
                "utm_medium":'',
                "utm_content":'',
                "utm_campaign":'',
                "campaign_id":'',
                "utm_adname":'',
                "adset_id":'',
                "fbclid":'',
                "ad_source":'',
                "ad_id":'',
                "university":'',
                "remarks":'',
                "remarks_timestamp":'',
                "fee_waiver_category":'',
                "total_questions":'',
                "document_status":'',
                "referral_code":'',
                "referred_code":'',
                "program":'',
                "reffered_by":''
            })

       
        lis.append({
                "name":"",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "fbc_id":'',
                "utm_source":'',
                "utm_medium":'',
                "utm_content":'',
                "utm_campaign":'',
                "campaign_id":'',
                "utm_adname":'',
                "adset_id":'',
                "fbclid":'',
                "ad_source":'',
                "ad_id":'',
                "university":'',
                "remarks":'',
                "remarks_timestamp":'',
                "fee_waiver_category":'',
                "total_questions":'',
                "document_status":'',
                "referral_code":'',
                "referred_code":'',
                "program":'',
                "reffered_by":''
            })
        
        lis.append({
                "name":"Full Name",
                "email":'Email',
                "subject":'Phone Number',
                "Chapter":'City',
                "Topic":'State',
                "fbc_id":'Fbc Id',
                "utm_source":'UTM Source',
                "utm_medium":'UTM Medium',
                "utm_content":'UTM Content',
                "utm_campaign":'UTM Campaign',
                "campaign_id":'Campaign Id',
                "utm_adname":'UTM Adname',
                "adset_id":'Adset Id',
                "fbclid":'Fbclid',
                "ad_source":'Ad Source',
                "ad_id":'Ad Id',
                "university":'University',
                "remarks":'Remarks',
                "remarks_timestamp":'Remarks Timestamp',
                "fee_waiver_category":'Fee Waiver Category',
                "total_questions":'Created At',
                "document_status":'Document Status',
                "referral_code":'Referral Code',
                "referred_code":'Referred Code',
                "program":'Program',
                "reffered_by":'Reffered By'
            })
        
        
        for chapter_data in serializers.data:
            lis.append({
                "name":chapter_data['full_name'],
                "email":chapter_data['email'],
                "subject":chapter_data['phone'],
                "Chapter":chapter_data['city'],
                "Topic":chapter_data['state'],
                "fbc_id":chapter_data['fbc_id'],
                "utm_source":chapter_data['utm_source'],
                "utm_medium":chapter_data['utm_medium'],
                "utm_content":chapter_data['utm_content'],
                "utm_campaign":chapter_data['utm_campaign'],
                "campaign_id":chapter_data['campaign_id'],
                "utm_adname":chapter_data['utm_adname'],
                "adset_id":chapter_data['adset_id'],
                "fbclid":chapter_data['fbclid'],
                "ad_source":chapter_data['ad_source'],
                "ad_id":chapter_data['ad_id'],
                "university":chapter_data['university'],
                "remarks":chapter_data['remarks'],
                "remarks_timestamp":chapter_data['remarks_timestamp'],
                "fee_waiver_category":chapter_data['fee_waiver_category'],
                "total_questions":chapter_data['created_at'],
                "document_status":chapter_data['document_status'],
                "referral_code":chapter_data['referral_code'],
                "referred_code":chapter_data['referred_code'],
                "program":chapter_data['program'],
                "reffered_by":chapter_data['reffered_by']
            })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "dossier_report"
            gcs_folder_name = "media/reports/dossier/excel"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)




############################ For Affiliates ##########################



class GetDossierSourceReportPDFView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]
    def get(self, request, sid=None):
        source_type = request.GET.get('source')
        if source_type:
            datas = DossierData.objects.filter(source=source_type).order_by('-id')
        else:
            datas = DossierData.objects.filter(source=SourceType.Website).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierDataSerializer(datas, many=True)


        data = {
                    "user_data":serializers.data,
                    "report_date": datetime.now().strftime("%d-%m-%Y, %H:%M")
                }
        

        template = get_template('pdf/dossier_report.html')
        html  = template.render(data)
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Encode HTML and create PDF
            html = html.encode('latin-1', 'replace').decode('latin-1')
            pdf = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=temp_file)

            if pdf.err:
                raise Exception("PDF generation error!")
        
        # After the 'with' block, the file is closed, but not deleted yet
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "dossier_report"
            gcs_folder_name = "media/source/excel"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.pdf"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted from the server's disk
            os.remove(pdf_path)
    

class GetDossierVSLSourceReportPDFView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","degree","degree_stage","phone"]
    ordering_fields = ['id',"full_name","email","phone","degree","degree_stage","created_at"]
    def get(self, request, sid=None):
        source_type = request.GET.get('source')
        if source_type:
            datas = DossierData.objects.filter(source=source_type).order_by('-id')
        else:
            datas = DossierData.objects.filter(source=SourceType.Website).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)

        degree = request.GET.get('degree')
        if degree:
            datas = datas.filter(degree__icontains=degree)

        degree_stage = request.GET.get('degree_stage')
        if degree_stage:
            datas = datas.filter(degree_stage__icontains=degree_stage)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierDataSerializer(datas, many=True)


        data = {
                    "user_data":serializers.data,
                    "report_date": datetime.now().strftime("%d-%m-%Y, %H:%M")
                }
        

        template = get_template('pdf/vsl_report.html')
        html  = template.render(data)
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Encode HTML and create PDF
            html = html.encode('latin-1', 'replace').decode('latin-1')
            pdf = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=temp_file)

            if pdf.err:
                raise Exception("PDF generation error!")
        
        # After the 'with' block, the file is closed, but not deleted yet
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "vsl_report"
            gcs_folder_name = "media/reports/vsl/pdf"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.pdf"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted from the server's disk
            os.remove(pdf_path)
    




class GetVSLAdvisorReportPDFView(APIView):
    # permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]
    def get(self, request, sid=None):

        adv_datas_list = list(VslDetail.objects.filter(specialist_status=True).values_list('dossier__id', flat=True))
        # print(adv_datas_list)
        datas = DossierData.objects.filter(id__in=adv_datas_list, source=SourceType.VslOptin).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierDataSerializer(datas, many=True)

        data = {
                    "user_data":serializers.data,
                    "report_date": datetime.now().strftime("%d-%m-%Y, %H:%M")
                }
        
        template = get_template('pdf/dossier_report.html')
        html  = template.render(data)
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Encode HTML and create PDF
            html = html.encode('latin-1', 'replace').decode('latin-1')
            pdf = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=temp_file)

            if pdf.err:
                raise Exception("PDF generation error!")
        
        # After the 'with' block, the file is closed, but not deleted yet
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "dossier_report"
            gcs_folder_name = "media/source/pdf"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.pdf"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted from the server's disk
            os.remove(pdf_path)




class GetDossierSourceReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]

    def get(self, request, sid=None):
        
        source_type = request.GET.get('source')
        if source_type:
            datas = DossierData.objects.filter(source=source_type).order_by('-id')
            if str(source_type) == str(SourceType.EAWebsite):
                access_data = settings.EAUTMSOURCE
                if str(request.user.first_name).lower() in access_data:
                    datas = DossierData.objects.filter(source=source_type, utm_source__iexact=str(request.user.first_name).lower()).order_by('-id')
            if str(source_type) == str(SourceType.CPAWebsite):
                access_data = settings.CPAUTMSOURCE
                if str(request.user.first_name).lower() in access_data:
                    datas = DossierData.objects.filter(source=source_type, utm_source__iexact=str(request.user.first_name).lower()).order_by('-id')

        else:
            datas = DossierData.objects.filter(source=SourceType.Website).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierDataSerializer(datas, many=True)

        lis = []
        if str(source_type) == "18":
            lis.append({
                    "name":"Full Name",
                    "email":'Email',
                    "subject":'Phone Number',
                    "Chapter":'City',
                    "Topic":'State',
                    "fbc_id":'Fbc Id',
                    "utm_source":'UTM Source',
                    "utm_medium":'UTM Medium',
                    "utm_content":'UTM Content',
                    "utm_campaign":'UTM Campaign',
                    "campaign_id":'Campaign Id',
                    "utm_adname":'UTM Adname',
                    "adset_id":'Adset Id',
                    "fbclid":'Fbclid',
                    "ad_source":'Ad Source',
                    "ad_id":'Ad Id',
                    "university":'University',
                    "remarks":'Remarks',
                    "remarks_timestamp":'Remarks Timestamp',
                    "fee_waiver_category":'Fee Waiver Category',
                    "total_questions":'Created At',
                    "document_status":'Document Status',
                    "referral_code":'Referral Code',
                    "referred_code":'Referred Code',
                    "program":'Program',
                    "reffered_by":'Referred By'
                })
            for chapter_data in serializers.data:
                lis.append({
                    "name":chapter_data['full_name'],
                    "email":chapter_data['email'],
                    "subject":chapter_data['phone'],
                    "Chapter":chapter_data['city'],
                    "Topic":chapter_data['state'],
                    "fbc_id":chapter_data['fbc_id'],
                    "utm_source":chapter_data['utm_source'],
                    "utm_medium":chapter_data['utm_medium'],
                    "utm_content":chapter_data['utm_content'],
                    "utm_campaign":chapter_data['utm_campaign'],
                    "campaign_id":chapter_data['campaign_id'],
                    "utm_adname":chapter_data['utm_adname'],
                    "adset_id":chapter_data['adset_id'],
                    "fbclid":chapter_data['fbclid'],
                    "ad_source":chapter_data['ad_source'],
                    "ad_id":chapter_data['ad_id'],
                    "university":chapter_data['university'],
                    "remarks":chapter_data['remarks'],
                    "remarks_timestamp":chapter_data['remarks_timestamp'],
                    "fee_waiver_category":chapter_data['fee_waiver_category'],
                    "total_questions":chapter_data['created_at'],
                    "document_status":chapter_data['document_status'],
                    "referral_code":chapter_data['referral_code'],
                    "referred_code":chapter_data['referred_code'],
                    "program":chapter_data['program'],
                    "reffered_by":chapter_data['reffered_by']
                })
        elif str(source_type) == "22":
            lis.append({
                    "name":"Full Name",
                    "email":'Email',
                    "subject":'Phone Number',
                    "Chapter":'City',
                    "Topic":'State',
                    "fbc_id":'Fbc Id',
                    "utm_source":'UTM Source',
                    "utm_medium":'UTM Medium',
                    "utm_content":'UTM Content',
                    "utm_campaign":'UTM Campaign',
                    "campaign_id":'Campaign Id',
                    "utm_adname":'UTM Adname',
                    "adset_id":'Adset Id',
                    "fbclid":'Fbclid',
                    "ad_source":'Ad Source',
                    "ad_id":'Ad Id',
                    "university":'University',
                    "remarks":'Remarks',
                    "remarks_timestamp":'Remarks Timestamp',
                    "fee_waiver_category":'Fee Waiver Category',
                    "total_questions":'Created At',
                    "document_status":'Document Status',
                    "referral_code":'Referral Code',
                    "referred_code":'Referred Code',
                    "program":'Program',
                    "reffered_by":'Referred By',
                    "interview_date":'Interview Date',
                    "age_range":'Age Range',
                    "fund_mode":'Fund Mode'
                })
            for chapter_data in serializers.data:
                lis.append({
                    "name":chapter_data['full_name'],
                    "email":chapter_data['email'],
                    "subject":chapter_data['phone'],
                    "Chapter":chapter_data['city'],
                    "Topic":chapter_data['state'],
                    "fbc_id":chapter_data['fbc_id'],
                    "utm_source":chapter_data['utm_source'],
                    "utm_medium":chapter_data['utm_medium'],
                    "utm_content":chapter_data['utm_content'],
                    "utm_campaign":chapter_data['utm_campaign'],
                    "campaign_id":chapter_data['campaign_id'],
                    "utm_adname":chapter_data['utm_adname'],
                    "adset_id":chapter_data['adset_id'],
                    "fbclid":chapter_data['fbclid'],
                    "ad_source":chapter_data['ad_source'],
                    "ad_id":chapter_data['ad_id'],
                    "university":chapter_data['university'],
                    "remarks":chapter_data['remarks'],
                    "remarks_timestamp":chapter_data['remarks_timestamp'],
                    "fee_waiver_category":chapter_data['fee_waiver_category'],
                    "total_questions":chapter_data['created_at'],
                    "document_status":chapter_data['document_status'],
                    "referral_code":chapter_data['referral_code'],
                    "referred_code":chapter_data['referred_code'],
                    "program":chapter_data['program'],
                    "reffered_by":chapter_data['reffered_by'],
                    "interview_date":chapter_data['interview_date'],
                    "age_range":chapter_data['age_range'],
                    "fund_mode":chapter_data['fund_mode']
                })
        elif str(source_type) == "23":
            lis.append({
                    "name":"Full Name",
                    "email":'Email',
                    "subject":'Phone Number',
                    "fbc_id":'Fbc Id',
                    "utm_source":'UTM Source',
                    "utm_medium":'UTM Medium',
                    "utm_content":'UTM Content',
                    "utm_campaign":'UTM Campaign',
                    "campaign_id":'Campaign Id',
                    "utm_adname":'UTM Adname',
                    "adset_id":'Adset Id',
                    "fbclid":'Fbclid',
                    "ad_source":'Ad Source',
                    "ad_id":'Ad Id',
                    "university":'University',
                    "remarks":'Remarks',
                    "remarks_timestamp":'Remarks Timestamp',
                    "fee_waiver_category":'Fee Waiver Category',
                    "total_questions":'Created At',
                    "document_status":'Document Status',
                    "referral_code":'Referral Code',
                    "referred_code":'Referred Code',
                    "program":'Program',
                    "reffered_by":'Referred By',
                    "interview_date":'Interview Date',
                    "age_range":'Age Range',
                    "fund_mode":'Fund Mode'
                })
            for chapter_data in serializers.data:
                lis.append({
                    "name":chapter_data['full_name'],
                    "email":chapter_data['email'],
                    "subject":chapter_data['phone'],
                    "Chapter":chapter_data['city'],
                    "Topic":chapter_data['state'],
                    "fbc_id":chapter_data['fbc_id'],
                    "utm_source":chapter_data['utm_source'],
                    "utm_medium":chapter_data['utm_medium'],
                    "utm_content":chapter_data['utm_content'],
                    "utm_campaign":chapter_data['utm_campaign'],
                    "campaign_id":chapter_data['campaign_id'],
                    "utm_adname":chapter_data['utm_adname'],
                    "adset_id":chapter_data['adset_id'],
                    "fbclid":chapter_data['fbclid'],
                    "ad_source":chapter_data['ad_source'],
                    "ad_id":chapter_data['ad_id'],
                    "university":chapter_data['university'],
                    "remarks":chapter_data['remarks'],
                    "remarks_timestamp":chapter_data['remarks_timestamp'],
                    "fee_waiver_category":chapter_data['fee_waiver_category'],
                    "total_questions":chapter_data['created_at'],
                    "document_status":chapter_data['document_status'],
                    "referral_code":chapter_data['referral_code'],
                    "referred_code":chapter_data['referred_code'],
                    "program":chapter_data['program'],
                    "reffered_by":chapter_data['reffered_by'],
                    "interview_date":chapter_data['interview_date'],
                    "age_range":chapter_data['age_range'],
                    "fund_mode":chapter_data['fund_mode']
                })
        else:
            lis.append({
                    "name":"Dossier Report",
                    "email":'',
                    "subject":'',
                    "Chapter":'',
                    "Topic":'',
                    "fbc_id":'',
                    "utm_source":'',
                    "utm_medium":'',
                    "utm_content":'',
                    "utm_campaign":'',
                    "campaign_id":'',
                    "utm_adname":'',
                    "adset_id":'',
                    "fbclid":'',
                    "ad_source":'',
                    "ad_id":'',
                    "university":'',
                    "remarks":'',
                    "remarks_timestamp":'',
                    "fee_waiver_category":'',
                    "total_questions":'',
                    "document_status":'',
                    "referral_code":'',
                    "referred_code":''
                })
            
        
            lis.append({
                    "name":"",
                    "email":'',
                    "subject":'',
                    "Chapter":'',
                    "Topic":'',
                    "fbc_id":'',
                    "utm_source":'',
                    "utm_medium":'',
                    "utm_content":'',
                    "utm_campaign":'',
                    "campaign_id":'',
                    "utm_adname":'',
                    "adset_id":'',
                    "fbclid":'',
                    "ad_source":'',
                    "ad_id":'',
                    "university":'',
                    "remarks":'',
                    "remarks_timestamp":'',
                    "fee_waiver_category":'',
                    "total_questions":'',
                    "document_status":'',
                    "referral_code":'',
                    "referred_code":''
                })
            
            lis.append({
                    "name":"Full Name",
                    "email":'Email',
                    "subject":'Phone Number',
                    "Chapter":'City',
                    "Topic":'State',
                    "fbc_id":'Fbc Id',
                    "utm_source":'UTM Source',
                    "utm_medium":'UTM Medium',
                    "utm_content":'UTM Content',
                    "utm_campaign":'UTM Campaign',
                    "campaign_id":'Campaign Id',
                    "utm_adname":'UTM Adname',
                    "adset_id":'Adset Id',
                    "fbclid":'Fbclid',
                    "ad_source":'Ad Source',
                    "ad_id":'Ad Id',
                    "university":'University',
                    "remarks":'Remarks',
                    "remarks_timestamp":'Remarks Timestamp',
                    "fee_waiver_category":'Fee Waiver Category',
                    "total_questions":'Created At',
                    "document_status":'Document Status',
                    "referral_code":'Referral Code',
                    "referred_code":'Referred Code'
                })
        
        
            for chapter_data in serializers.data:
                lis.append({
                    "name":chapter_data['full_name'],
                    "email":chapter_data['email'],
                    "subject":chapter_data['phone'],
                    "Chapter":chapter_data['city'],
                    "Topic":chapter_data['state'],
                    "fbc_id":chapter_data['fbc_id'],
                    "utm_source":chapter_data['utm_source'],
                    "utm_medium":chapter_data['utm_medium'],
                    "utm_content":chapter_data['utm_content'],
                    "utm_campaign":chapter_data['utm_campaign'],
                    "campaign_id":chapter_data['campaign_id'],
                    "utm_adname":chapter_data['utm_adname'],
                    "adset_id":chapter_data['adset_id'],
                    "fbclid":chapter_data['fbclid'],
                    "ad_source":chapter_data['ad_source'],
                    "ad_id":chapter_data['ad_id'],
                    "university":chapter_data['university'],
                    "remarks":chapter_data['remarks'],
                    "remarks_timestamp":chapter_data['remarks_timestamp'],
                    "fee_waiver_category":chapter_data['fee_waiver_category'],
                    "total_questions":chapter_data['created_at'],
                    "document_status":chapter_data['document_status'],
                    "referral_code":chapter_data['referral_code'],
                    "referred_code":chapter_data['referred_code']
                })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "dossier_report"
            gcs_folder_name = "media/reports/source/excel"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)

class GetVSLAdvisorReportExcelView(APIView):
    # permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]

    def get(self, request, sid=None):
        
        adv_datas_list = list(VslDetail.objects.filter(specialist_status=True).values_list('dossier__id', flat=True))
        # print(adv_datas_list)
        datas = DossierData.objects.filter(id__in=adv_datas_list, source=SourceType.VslOptin).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierDataSerializer(datas, many=True)

        lis = []
        
        lis.append({
                "name":"Dossier Report",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "fbc_id":'',
                "utm_source":'',
                "utm_medium":'',
                "utm_content":'',
                "utm_campaign":'',
                "campaign_id":'',
                "utm_adname":'',
                "adset_id":'',
                "fbclid":'',
                "ad_source":'',
                "ad_id":'',
                "university":'',
                "remarks":'',
                "remarks_timestamp":'',
                "fee_waiver_category":'',
                "total_questions":''
            })

       
        lis.append({
                "name":"",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "fbc_id":'',
                "utm_source":'',
                "utm_medium":'',
                "utm_content":'',
                "utm_campaign":'',
                "campaign_id":'',
                "utm_adname":'',
                "adset_id":'',
                "fbclid":'',
                "ad_source":'',
                "ad_id":'',
                "university":'',
                "remarks":'',
                "remarks_timestamp":'',
                "fee_waiver_category":'',
                "total_questions":''
            })
        
        lis.append({
                "name":"Full Name",
                "email":'Email',
                "subject":'Phone Number',
                "Chapter":'City',
                "Topic":'State',
                "fbc_id":'Fbc Id',
                "utm_source":'UTM Source',
                "utm_medium":'UTM Medium',
                "utm_content":'UTM Content',
                "utm_campaign":'UTM Campaign',
                "campaign_id":'Campaign Id',
                "utm_adname":'UTM Adname',
                "adset_id":'Adset Id',
                "fbclid":'Fbclid',
                "ad_source":'Ad Source',
                "ad_id":'Ad Id',
                "university":'University',
                "remarks":'Remarks',
                "remarks_timestamp":'Remarks Timestamp',
                "fee_waiver_category":'Fee Waiver Category',
                "total_questions":'Created At'
            })
        
        
        for chapter_data in serializers.data:
            lis.append({
                "name":chapter_data['full_name'],
                "email":chapter_data['email'],
                "subject":chapter_data['phone'],
                "Chapter":chapter_data['city'],
                "Topic":chapter_data['state'],
                "fbc_id":chapter_data['fbc_id'],
                "utm_source":chapter_data['utm_source'],
                "utm_medium":chapter_data['utm_medium'],
                "utm_content":chapter_data['utm_content'],
                "utm_campaign":chapter_data['utm_campaign'],
                "campaign_id":chapter_data['campaign_id'],
                "utm_adname":chapter_data['utm_adname'],
                "adset_id":chapter_data['adset_id'],
                "fbclid":chapter_data['fbclid'],
                "ad_source":chapter_data['ad_source'],
                "ad_id":chapter_data['ad_id'],
                "university":chapter_data['university'],
                "remarks":chapter_data['remarks'],
                "remarks_timestamp":chapter_data['remarks_timestamp'],
                "fee_waiver_category":chapter_data['fee_waiver_category'],
                "total_questions":chapter_data['created_at']
            })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "dossier_report"
            gcs_folder_name = "media/reports/source/excel"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)

class GetAmendmentSourceReportExcelView(APIView):
    # permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone"]
    ordering_fields = ['id',"full_name","email","phone","created_at"]

    def get(self, request, sid=None):
        
        source_type = request.GET.get('source')
        if source_type:
            datas = DossierAbondant.objects.filter(source=source_type).order_by('-id')
        else:
            datas = DossierAbondant.objects.filter(source=SourceType.Website).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierAbondantSerializer(datas, many=True)

        lis = []
        
        lis.append({
                "full_name":"Amendment Report",
                "email":'',
                "phone":'',
                "utm_source":'',
                "utm_medium":'',
                "utm_campaign":'',
                "created_at":''
            })

       
        lis.append({
                "full_name":"",
                "email":'',
                "phone":'',
                "utm_source":'',
                "utm_medium":'',
                "utm_campaign":'',
                "created_at":''
            })
        
        lis.append({
                "full_name":"Full Name",
                "email":'Email',
                "phone":'Phone Number',
                "utm_source":'UTM Source',
                "utm_medium":'UTM Medium',
                "utm_campaign":'UTM Campaign',
                "created_at":'Date Time'
            })
        
        
        for chapter_data in serializers.data:
            lis.append({
                "full_name":chapter_data["full_name"],
                "email":chapter_data["email"],
                "phone":chapter_data["phone"],
                "utm_source":chapter_data["utm_source"],
                "utm_medium":chapter_data["utm_medium"],
                "utm_campaign":chapter_data["utm_campaign"],
                "created_at":chapter_data["created_at"]
            })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "amendment_report"
            gcs_folder_name = "media/reports/source/excel"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)



class GetDossierVSLSourceReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]

    def get(self, request, sid=None):
        
        source_type = request.GET.get('source')
        if source_type:
            datas = DossierData.objects.filter(source=source_type).order_by('-id')
        else:
            datas = DossierData.objects.filter(source=SourceType.Website).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        degree = request.GET.get('degree')
        if degree:
            datas = datas.filter(degree__icontains=degree)

        degree_stage = request.GET.get('degree_stage')
        if degree_stage:
            datas = datas.filter(degree_stage__icontains=degree_stage)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierDataSerializer(datas, many=True)

        lis = []
        
        lis.append({
                "name":"VSL Report",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "total_questions":''
            })

       
        lis.append({
                "name":"",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "total_questions":''
            })
        
        lis.append({
                "name":"Full Name",
                "email":'Email',
                "subject":'Phone Number',
                "Chapter":'Degree',
                "Topic":'Degree Stage',
                "total_questions":'Created At',
            })
        
        
        for chapter_data in serializers.data:
            lis.append({
                "name":chapter_data['full_name'],
                "email":chapter_data['email'],
                "subject":chapter_data['phone'],
                "Chapter":chapter_data['degree'],
                "Topic":chapter_data['degree_stage'],
                "total_questions":chapter_data['created_at'],
            })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "vsl_report"
            gcs_folder_name = "media/reports/vsl/excel"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)





class GetDeleteLead(APIView):
    def get(self, request, sid=None):
        print("deleting.....")
        # d_list = [257,694,727,744,748,750,752,773,774,782,894,254,503,1147,3561,3795,3524,3780,3526,4109,1171,1270,4109]
        # d_list = [284]
        # d_list = [519,313,506,404,387,489,505,509]
        # d_list = [5562,5561,5403,4547,4554,4553,4509,4508,4506,427,426,425]
        # d_list = [6117,2328,3377,6376,1013,6373,6375,248,3268,4450,4614,6337,251,272,961,962,2823,3610,3807,4341,4628,5555,5724,5899,6374,6372,6370,6366,6362,5315,6379,6377,6500,6499,6498,6513,6931,6536,6554,6471]
        d_list = []
        for i in d_list:
            d = DossierData.objects.filter(id=i)
            if d:
                d.first().delete()
                print(i,">>>>>>deleted....")
        return success_response(
                message="Success",
                data={},
                status_code=status.HTTP_200_OK
            )



















###################### Extra ###################
from openpyxl import Workbook
from openpyxl.styles import Font
# from django.http import HttpResponse
import pandas as pd
from io import BytesIO
from django.http import HttpResponse
import uuid

# class ExcelLogicProcessAPI(APIView):

#     def post(self, request, *args, **kwargs):
#         print("starting...")
#         file = request.FILES.get("file")
#         print("starting...1")

#         if not file:
#             return Response(
#                 {"error": "Excel file is required"},
#                 status=status.HTTP_400_BAD_REQUEST
#             )
#         print("dd")
#         try:
#             # =========================
#             # READ EXCEL FILE
#             # =========================
#             df = pd.read_csv(file)
#             # print(df)
#             # =========================
#             # VALIDATE REQUIRED COLUMN
#             # =========================
#             required_columns = [
#                 "Registered Name",
#                 "Registered Email",
#                 "Registered Mobile"
#             ]

#             print("starting...2")
#             missing_columns = [
#                 col for col in required_columns
#                 if col not in df.columns
#             ]

#             if missing_columns:
#                 return Response(
#                     {
#                         "error": f"Missing columns: {missing_columns}"
#                     },
#                     status=400
#                 )

#             # =========================
#             # CLEAN DATA
#             # =========================
#             df["Registered Mobile"] = (
#                 df["Registered Mobile"]
#                 .astype(str)
#                 .str.replace("+91", "", regex=False)
#                 .str.replace("-", "", regex=False)
#                 .str.replace(" ", "", regex=False)
#                 .str.replace("`", "", regex=False)
#                 .str.strip()
#             )

#             df["Registered Email"] = (
#                 df["Registered Email"]
#                 .astype(str)
#                 .str.lower()
#                 .str.strip()
#             )

#             # =========================
#             # REMOVE DUPLICATES
#             # =========================
#             df.drop_duplicates(
#                 subset=["Registered Mobile"],
#                 keep="first",
#                 inplace=True
#             )
#             print(df)
#             # =========================
#             # LOGIC PROCESS
#             # =========================
#             updated_data = []

#             for _, row in df.iterrows():

#                 mobile = row["Registered Mobile"]
#                 email = row["Registered Email"]
#                 print(mobile)
#                 print(email)



#             #     # Match with database
#             #     lead = DossierData.objects.filter(
#             #         mobile__icontains=mobile
#             #     ).first()

#             #     if lead:
#             #         status_value = "EXISTING LEAD"
#             #         db_email = lead.email
#             #     else:
#             #         status_value = "NEW LEAD"
#             #         db_email = ""

#                 updated_data.append({
#                     "Registered Name": row["Registered Name"],
#                     "Registered Email": email,
#                     "Registered Mobile": mobile
#                 })

#             # =========================
#             # CREATE OUTPUT EXCEL
#             # =========================
#             wb = Workbook()
#             ws = wb.active
#             ws.title = "Processed Data"

#             headers = [
#                 "Registered Name",
#                 "Registered Email",
#                 "Registered Mobile"
#             ]

#             # Add Header
#             for col_num, header in enumerate(headers, 1):
#                 cell = ws.cell(row=1, column=col_num)
#                 cell.value = header
#                 cell.font = Font(bold=True)

#             # Add Rows
#             for row_num, item in enumerate(updated_data, 2):
#                 ws.cell(row=row_num, column=1).value = item["Registered Name"]
#                 ws.cell(row=row_num, column=2).value = item["Registered Email"]
#                 ws.cell(row=row_num, column=3).value = item["Registered Mobile"]

#             # =========================
#             # RETURN EXCEL RESPONSE
#             # =========================
#             output = BytesIO()
#             wb.save(output)
#             output.seek(0)

#             with tempfile.NamedTemporaryFile(
#                 suffix='.xlsx',
#                 delete=False
#             ) as temp_file:

#                 excel_path = temp_file.name

#                 output_df = pd.DataFrame(updated_data)

#                 output_df.to_excel(
#                     excel_path,
#                     index=False
#                 )

#             try:

#                 # =========================
#                 # GCP STORAGE UPLOAD
#                 # =========================

#                 timestamp = datetime.now().strftime(
#                     "%d_%m_%Y_%H_%M_%S"
#                 )

#                 report_name = "processed_leads"

#                 gcs_folder_name = (
#                     "media/reports/excel"
#                 )

#                 gcs_file_name = (
#                     f"{gcs_folder_name}/"
#                     f"{report_name}_{timestamp}.xlsx"
#                 )

#                 # Initialize GCP Client
#                 client = storage.Client()

#                 # Bucket
#                 bucket = client.get_bucket(
#                     settings.GS_BUCKET_NAME
#                 )

#                 # Blob
#                 blob = bucket.blob(gcs_file_name)

#                 # Upload File
#                 blob.upload_from_filename(excel_path)

#                 # Public URL
#                 file_url = blob.public_url

#                 # =========================
#                 # RESPONSE
#                 # =========================
#                 return Response(
#                     {
#                         "status": True,
#                         "message": "Excel processed successfully",
#                         "file_url": file_url
#                     },
#                     status=status.HTTP_200_OK
#                 )

#             finally:

#                 # Delete Temp File
#                 if os.path.exists(excel_path):
#                     os.remove(excel_path)

#         except Exception as e:
#             return Response(
#                 {"error": str(e)},
#                 status=500
#             )



import os
import pandas as pd
import tempfile
from datetime import datetime
from google.cloud import storage
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings


class ExcelLogicProcessAPI(APIView):

    def post(self, request, *args, **kwargs):

        try:

            file = request.FILES.get("file")

            if not file:
                return Response(
                    {"error": "Excel file is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # =========================
            # READ FILE
            # =========================
            file_name = file.name.lower()

            if file_name.endswith(".csv"):
                df = pd.read_csv(file)

            elif file_name.endswith(".xlsx"):
                df = pd.read_excel(file, engine="openpyxl")

            else:
                return Response(
                    {"error": "Only CSV/XLSX files allowed"},
                    status=400
                )

            # =========================
            # VALIDATE REQUIRED COLUMNS
            # =========================
            required_columns = [
                "Registered Name",
                "Registered Email",
                "Registered Mobile"
            ]

            missing_columns = [
                col for col in required_columns
                if col not in df.columns
            ]

            if missing_columns:
                return Response(
                    {
                        "error": f"Missing columns: {missing_columns}"
                    },
                    status=400
                )

            # =========================
            # CLEAN DATA
            # =========================
            df["Registered Mobile"] = (
                df["Registered Mobile"]
                .astype(str)
                .str.replace("+91", "", regex=False)
                .str.replace("-", "", regex=False)
                .str.replace(" ", "", regex=False)
                .str.replace("`", "", regex=False)
                .str.strip()
            )

            df["Registered Email"] = (
                df["Registered Email"]
                .astype(str)
                .str.lower()
                .str.strip()
            )

            # =========================
            # REMOVE DUPLICATES
            # =========================
            df.drop_duplicates(
                subset=["Registered Mobile"],
                keep="first",
                inplace=True
            )

            # =========================
            # APPLY LOGIC
            # =========================
            updated_data = []

            for _, row in df.iterrows():

                mobile = row["Registered Mobile"]
                email = row["Registered Email"]
                # if mobile and email:
                print(mobile, email)
                if (str(mobile).upper() != "NA") & (str(email).upper() != "NA"):
                    
                # =========================
                # YOUR DB LOGIC
                # =========================

                # lead = DossierData.objects.filter(
                #     mobile__icontains=mobile
                # ).first()

                # if lead:
                #     lead_status = "EXISTING"
                # else:
                #     lead_status = "NEW"

                    updated_data.append({
                        "Registered Name": row["Registered Name"],
                        "Registered Email": email,
                        "Registered Mobile": mobile,
                    })

            # =========================
            # CREATE TEMP EXCEL FILE
            # =========================
            with tempfile.NamedTemporaryFile(
                suffix='.xlsx',
                delete=False
            ) as temp_file:

                excel_path = temp_file.name

                output_df = pd.DataFrame(updated_data)

                output_df.to_excel(
                    excel_path,
                    index=False
                )

            try:

                # =========================
                # GCP STORAGE UPLOAD
                # =========================

                timestamp = datetime.now().strftime(
                    "%d_%m_%Y_%H_%M_%S"
                )

                report_name = "processed_leads"

                gcs_folder_name = (
                    "media/reports/excel"
                )

                gcs_file_name = (
                    f"{gcs_folder_name}/"
                    f"{report_name}_{timestamp}.xlsx"
                )

                # Initialize GCP Client
                client = storage.Client()

                # Bucket
                bucket = client.get_bucket(
                    settings.GS_BUCKET_NAME
                )

                # Blob
                blob = bucket.blob(gcs_file_name)

                # Upload File
                blob.upload_from_filename(excel_path)

                # Public URL
                file_url = blob.public_url

                # =========================
                # RESPONSE
                # =========================
                return Response(
                    {
                        "status": True,
                        "message": "Excel processed successfully",
                        "file_url": file_url
                    },
                    status=status.HTTP_200_OK
                )

            finally:

                # Delete Temp File
                if os.path.exists(excel_path):
                    os.remove(excel_path)

        except Exception as e:

            return Response(
                {
                    "status": False,
                    "error": str(e)
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )




##########################################################################################################
### testing ###

import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

class ImportEmailView(APIView):

    def post(self, request):
        excel_file = request.FILES.get("file")

        if not excel_file:
            return Response(
                {"message": "Please upload an Excel file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Read Excel
            # df = pd.read_excel(excel_file)
            df = pd.read_csv(excel_file)

            # Check column exists
            if "Registered Email" not in df.columns:
                return Response(
                    {"message": "Registered Email column not found."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get emails
            emails = (
                df["Registered Email"]
                .dropna()
                .astype(str)
                .str.strip()
                .tolist()
            )

            # # print(emails)
            # j = 0
            # # all_data = list(DossierData.objects.all().values_list('email', flat=True))
            # all_data = list(User.objects.all().values('email','application_id'))
            # # all_data = User.objects.all().values('email','application_id')
            # for i in emails:
            #     for k in all_data:
            #         if i.lower() == str(k["email"]).lower():
            #             print(k)
            #             j+=1
                        
            #             url = settings.MERITO_BASE_URL+"/lead/v1/createOrUpdate"
            #             headers = {
            #                 "Content-Type": "application/json",
            #                 "secret-key": settings.MERITO_SECRETE_KEY,
            #                 "access-key": settings.MERITO_ACCESS_KEY
            #             }
            #             payload = {
            #                 "email": i.lower(),
            #                 "search_criteria": "email",
            #                 "cf_gcc_application_number":k["application_id"]
            #             }
            #             try:
            #                 response = requests.post(url, headers=headers, json=payload)
            #                 print(response.status_code)
            #                 print(response.text)
            #             except Exception as e:
            #                 print("API Error:", str(e))

            #         # j+=1

            # print(j)

            headers = {
                "Content-Type": "application/json",
                "secret-key": settings.MERITO_SECRETE_KEY,
                "access-key": settings.MERITO_ACCESS_KEY
            }

            # all_users = {
            #         str(user["email"]).lower(): user["application_id"]
            #         for user in User.objects.exclude(email__isnull=True)
            #         .values("email", "application_id")
            #     }
            all_data = list(DossierData.objects.all().values_list('email', flat=True))
            matched_count = 0
            l = 0
            for email in emails:
                # ee = all_users.get(email.lower())
                # print(type(ee))
                # if str(email).lower() in all_data:
                payload = {
                    "email": str(email).lower(),
                    "search_criteria": "email",
                    "cf_payment_status":"Pending"
                }

                try:
                    response = requests.post(
                        settings.MERITO_BASE_URL + "/lead/v1/createOrUpdate",
                        headers=headers,
                        json=payload,
                        timeout=30
                    )

                    print(email, response.status_code)
                    l+=1
                except Exception as e:
                    print(f"{email}: {str(e)}")

                matched_count+=1
                print(matched_count)
                
            return Response({
                "message": "Success",
                "total_emails": len(emails),
                "total_website_emails": matched_count,
                "l":l,
                "emails": emails
            })

        except Exception as e:
            return Response(
                {"message": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


### for affliate 6 data export ##############


class GetDossierAffliateSixReportExcelView(APIView):
    def get(self, request, sid=None):
        # datas = DossierData.objects.filter(source=SourceType.Affiliate6).order_by('-id')
        datas = (DossierData.objects.filter(source=SourceType.Affiliate6).order_by('phone', '-id').distinct('phone'))
        # print(len(datas))
        # return Response({})
        data_list = ListDossierDataAffliateSixReportSerializer(datas, many=True).data
        COLUMN_MAPPING = {
            "full_name":'Full Name',
            "email":'Email',
            "phone":'Phone',
            "city":'City',
            "state":'State',
            "fbc_id":'FBC ID',
            "utm_source":'UTM SOURCE',
            "utm_medium":'UTM MEDIUM',
            "utm_content":'UTM CONTENT',
            "utm_campaign":'UTM CAMPAIGN',
            "campaign_id":'CAMPAIGN ID',
            "utm_adname":'UTM ADNAME',
            "adset_id":'ADSET ID',
            "fbclid":'FBCLID',
            "ad_source":'AD SOURCE',
            "ad_id":'AD ID',
            "fee_waiver_category":'FEE WAIVER CATEGORY',
            "created_at":'CREATE TIMESTAMP'
        }
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name

            df = pd.DataFrame(data_list)
            # df = df.drop_duplicates(
            #     subset=['phone'],
            #     keep='first'
            # )
            df = df[list(COLUMN_MAPPING.keys())]

            df.rename(columns=COLUMN_MAPPING, inplace=True)

            df.to_excel(
                pdf_path,
                header=True,
                index=False
            )
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "lead_report"
            gcs_folder_name = "media/reports/dossier/excel/affliatesix"
            gcs_file_name = f"{gcs_folder_name}/{report_name}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            # overwrite existing file
            # if blob.exists():
            #     blob.delete(igno)
                # blob.upload_from_filename(pdf_path)

            # upload new excel
            # blob.upload_from_filename(
            #     pdf_path,
            #     content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            # )
            # blob.reload()
            # print(blob.updated)
            # print("updated:", blob.updated)
            # print("generation:", blob.generation)
            blob = bucket.blob(gcs_file_name)

            blob.upload_from_filename(
                pdf_path,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            blob.cache_control = "no-cache"
            blob.patch()

            print(blob.updated)
            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)




import pandas as pd
from django.http import HttpResponse


import pandas as pd
from django.http import HttpResponse
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


class GetDossierAffliateSixReportExcelView(APIView):

    def get(self, request, sid=None):

        datas = (
            DossierData.objects
            .filter(source=SourceType.Affiliate6)
            .order_by('id')
        )

        # remove duplicate phone
        seen = set()
        unique_data = []

        for item in datas:
            if item.phone not in seen:
                seen.add(item.phone)
                unique_data.append(item)


        data_list = ListDossierDataAffliateSixReportSerializer(
            unique_data,
            many=True
        ).data


        COLUMN_MAPPING = {
            "full_name": 'Full Name',
            "email": 'Email',
            "phone": 'Phone',
            "city": 'City',
            "state": 'State',
            "fbc_id": 'FBC ID',
            "utm_source": 'UTM SOURCE',
            "utm_medium": 'UTM MEDIUM',
            "utm_content": 'UTM CONTENT',
            "utm_campaign": 'UTM CAMPAIGN',
            "campaign_id": 'CAMPAIGN ID',
            "utm_adname": 'UTM ADNAME',
            "adset_id": 'ADSET ID',
            "fbclid": 'FBCLID',
            "ad_source": 'AD SOURCE',
            "ad_id": 'AD ID',
            "fee_waiver_category": 'FEE WAIVER CATEGORY',
            "created_at": 'CREATE TIMESTAMP'
        }


        df = pd.DataFrame(data_list)


        df = df[list(COLUMN_MAPPING.keys())]


        df.rename(
            columns=COLUMN_MAPPING,
            inplace=True
        )


        # create excel in memory
        excel_file = BytesIO()

        df.to_excel(
            excel_file,
            index=False
        )

        excel_file.seek(0)


        # format excel
        wb = load_workbook(excel_file)
        ws = wb.active


        # header style
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center"
            )


        # auto width
        for column in ws.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            ws.column_dimensions[column_letter].width = max_length + 5


        response = HttpResponse(
            content_type=
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="lead_report.xlsx"'
        )


        wb.save(response)


        return response



class GetDossierInterviewAffliateSevenReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["full_name","email","phone","city","state","interview_date"]
    ordering_fields = ['id']

    def get(self, request, sid=None):
        
        datas = DossierData.objects.filter(source=SourceType.Affiliate7).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListDossierDataAffliateSevenInterviewSerializer(datas, many=True)

        lis = []
        
        lis.append({
                "full_name":"Dossier Interview Report",
                "email":'',
                "phone":'',
                "interview_date":'',
                "created_at":''
            })

       
        lis.append({
                "full_name":"",
                "email":'',
                "phone":'',
                "interview_date":'',
                "created_at":''
            })
        
        lis.append({
                "full_name":"Full Name",
                "email":'Email',
                "phone":'Phone Number',
                "interview_date":'Interview Date',
                "created_at":'Register Date&Time'
            })
        
        
        for chapter_data in serializers.data:
            lis.append({
                "full_name":chapter_data['full_name'],
                "email":chapter_data['email'],
                "phone":chapter_data['phone'],
                "interview_date":chapter_data['interview_date'],
                "created_at":chapter_data['created_at'],
            })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "dossier_interview_report"
            gcs_folder_name = "media/reports/dossier/excel"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)








##########################

from utils.google_sheet import get_google_sheet


class GetAffliateSixExcelView(APIView):
    def post(self, request, sid=None):

        sheet = get_google_sheet()

        # row = [
        #     dossier.get("full_name"),
        #     dossier.get("email"),
        #     dossier.get("phone"),
        #     dossier.get("city"),
        #     dossier.get("state"),
        #     dossier.get("created_at"),
        # ]
        {
            "full_name": 'Full Name',
            "email": 'Email',
            "phone": 'Phone',
            "city": 'City',
            "state": 'State',
            "fbc_id": 'FBC ID',
            "utm_source": 'UTM SOURCE',
            "utm_medium": 'UTM MEDIUM',
            "utm_content": 'UTM CONTENT',
            "utm_campaign": 'UTM CAMPAIGN',
            "campaign_id": 'CAMPAIGN ID',
            "utm_adname": 'UTM ADNAME',
            "adset_id": 'ADSET ID',
            "fbclid": 'FBCLID',
            "ad_source": 'AD SOURCE',
            "ad_id": 'AD ID',
            "fee_waiver_category": 'FEE WAIVER CATEGORY',
            "created_at": 'CREATE TIMESTAMP'
        }
        # row = [
        #     obj.full_name,
        #     obj.email,
        #     obj.phone,
        #     obj.city,
        #     obj.state,
        #     obj.fbc_id,
        #     obj.utm_source,
        #     obj.utm_medium,
        #     obj.utm_content,
        #     obj.utm_campaign,
        #     obj.campaign_id,
        #     obj.utm_adname,
        #     obj.adset_id,
        #     obj.fbclid,
        #     obj.ad_source,
        #     obj.ad_id,
        #     obj.fee_waiver_category,
        #     obj.created_at
        # ]
        
        

        # sheet.append_row(row)


        return success_response(
                message="Success",
                data={"report_url": ""},
                status_code=status.HTTP_200_OK
            )


class GetAffliateSevenLeadAllReportExcelView(APIView):
    # permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"full_name","email","phone","city","state"]
    ordering_fields = ['id',"full_name","email","phone","city","state","created_at"]

    def get(self, request, sid=None):
        
        datas = DossierData.objects.filter(source=SourceType.Affiliate7).order_by('id')
        # remove duplicate phone
        seen = set()
        unique_data = []

        for item in datas:
            if item.phone not in seen:
                seen.add(item.phone)
                unique_data.append(item)

        serializers = ListDossierDataAffliateSevenInterviewLiveReportSerializer(unique_data, many=True)

        lis = []
        
        lis.append({
                "full_name":"Dossier Report",
                "email":'',
                "phone":'',
                "city":'',
                "state":'',
                "interview_booked_status":'',
                "interview_date":'',
                "created_at":''
            })

       
        lis.append({
                "full_name":"Dossier Report",
                "email":'',
                "phone":'',
                "city":'',
                "state":'',
                "interview_booked_status":'',
                "interview_date":'',
                "created_at":''
            })
        
        lis.append({
                "full_name":"full name",
                "email":'Email',
                "phone":'Phone',
                "city":'City',
                "state":'State',
                "interview_booked_status":'status',
                "interview_date":'interview date',
                "created_at":'Created_at'
            })
        
        
        for chapter_data in serializers.data:
            lis.append({
                "full_name":chapter_data['full_name'],
                "email":chapter_data['email'],
                "phone":chapter_data['phone'],
                "city":chapter_data['city'],
                "state":chapter_data['state'],
                "interview_booked_status":chapter_data['interview_booked_status'],
                "interview_date":chapter_data['interview_date'],
                "created_at":chapter_data['created_at']
            })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "dossier_report"
            gcs_folder_name = "media/reports/source/excel"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)




def generate_time_slots(date_str, speak_with):
    start = datetime.strptime(f"{date_str} 10:00", "%Y-%m-%d %H:%M")
    end = datetime.strptime(f"{date_str} 21:00", "%Y-%m-%d %H:%M")

    # Booked slots
    booked_slots = set(
        DossierData.objects.filter(
            source=23,
            interview_date=date_str,
            speak_with=speak_with
        ).values_list("slot_time", flat=True)
    )

    now = timezone.localtime()
    today_str = now.strftime("%Y-%m-%d")
    cutoff_time = now + timedelta(minutes=90)

    slots = []
    all_slots = []

    while start < end:
        slot_end = start + timedelta(minutes=45)

        if slot_end > end:
            break

        # Skip past slots if selected date is today
        if date_str == today_str:
            slot_datetime = timezone.make_aware(start)
            if slot_datetime <= cutoff_time:
                start = slot_end
                continue

        time_str = start.strftime("%I:%M %p")

        slots.append({
            "start_time": time_str,
            "end_time": slot_end.strftime("%I:%M %p"),
            "book_status": 1 if (time_str in booked_slots) or (time_str in ["01:45 PM","02:30 PM"]) else 0
        })
        all_slots.append({
            "start_time": time_str,
            "end_time": slot_end.strftime("%I:%M %p"),
            # "book_status": 1 if (time_str in booked_slots) or (time_str in ["01:45 PM","02:30 PM"]) else 0
        })

        start = slot_end

    return slots, all_slots



class DossierTimeSlotAPIView(APIView):
    def get(self, request):
        date = request.query_params.get("date")  # 2026-08-14
        speak_with = request.query_params.get("speak_with")  # 1/2

        if not date:
            return Response({"error": "date is required"}, status=400)
        if not speak_with:
            return Response({"error": "speak_with is required"}, status=400)

        slots, all_slots = generate_time_slots(date, speak_with)

        return Response({
            "date": date,
            "slots": slots,
            "all_slots": all_slots
        })



class RescheduleInviteView(APIView):
    def post(self, request):
        id= request.data.get("lead_id")
        int_date = request.data.get("selected_date")
        int_time = request.data.get("selected_time")
        speak_with = request.data.get("speak_with")
        obj = DossierData.objects.get(id=id)
        obj.slot_time = int_time
        obj.interview_date = int_date
        obj.speak_with = speak_with
        obj.save()
        objs = ManageDossierMeeting.objects.filter(dossier_id=obj.id)
        if objs:
            objs = objs.last()
            try:
                zoom.cancel_zoom_meeting(objs.meeting_id)
            except:
                pass
        if settings.EXCEL_INPUT == "True":
            print("email sending start")

            # resend_email_invite(obj.id)

            threading.Thread(
                target=resend_email_invite,
                args=(str(obj.id),),
                daemon=True,
            ).start()

        return success_response(
            message="Success",
            data={},
            status_code=status.HTTP_200_OK
        )
    
    def get(self, request):
        from datetime import datetime
        from utils.google_meet import create_google_meet
        start = datetime.fromisoformat("2026-08-20T11:00:00")

        meeting = create_google_meet(
            topic="GCC Session",
            start_time=start,
            duration=45,
            attendees=["mukulsoft@gmail.com","vkd2695@gmail.com"]
        )
        print(meeting)
        return Response({
                    "date": ""
            })
    
    def delete(self, request):
        id= request.data.get("lead_id")
        DossierData.objects.filter(id=id).update(slot_time="")
        obj = ManageDossierMeeting.objects.filter(dossier_id=id)
        if obj:
            if settings.EXCEL_INPUT == "True":
                datas = obj.last()
                meeting_id = datas.meeting_id
                print("email sending start")
                try:
                    statuss = zoom.cancel_zoom_meeting(meeting_id)
                    print("cancel response....",statuss)
                    datas.cancel_status = True
                    datas.save()
                except requests.exceptions.HTTPError as e:
                    status_code = e.response.status_code if e.response else 502
                    if status_code == 404:
                        return Response(
                            {"error": "Meeting not found or already deleted."},
                            status=status.HTTP_404_NOT_FOUND,
                        )
                    return Response(
                        {"error": f"Failed to cancel meeting: {str(e)}"},
                        status=status.HTTP_502_BAD_GATEWAY,
                    )
            return success_response(
                message="Success",
                data={},
                status_code=status.HTTP_200_OK
            )
        return Response({
            "message":"No Data Found",
            "data":{},
            "status":400,
        })


phone_key= [
    "7027226387",
    "9123022270",
    "8960047463",
    "9920915132",
    "7874785748",
    "7054240215",
    "9667583222",
    "8596859685",
    "7974832505",
    "8978641413",
    "6398487985",
    "9465410327",
    "9560174404",
    "8218713851",
    "9306672983",
    "7988291633",
    "9149152056",
    "7027321020",
    "8423872376",
    "7217742531",
    "9811750022",
    "7078994558",
    "8851565977",
    "9818983282",
    "8685096604",
    "7678695103",
    "7302113094",
    "9690653416",
    "8882128533",
    "9336547682",
    "9811646475",
    "8580665265",
    "8209286224",
    "6282559482",
    "6367374242",
    "9243436689",
    "8475923868",
    "7389403296",
    "7483124898",
    "8053873843",
    "8279816735",
    "8009161462",
    "7307464231",
    "9801590312",
    "7014093946",
    "9258704713",
    "8954623994",
    "8097183653",
    "6204607833",
    "9812435176",
    "7055573932",
    "7668714851",
    "9236134969",
    "7079271723",
    "9639473170",
    "9762061037",
    "7302662607",
    "8745020559",
    "6395548992",
    "7303677149",
    "9958633068",
    "9253284392",
    "7905381439",
    "9339371908",
    "8287908357",
    "9334241013",
    "8755968541",
    "8510808131",
    "7805009614",
    "8923888322",
    "7080020288",
    "8448324625",
    "9729038764",
    "9999857015",
    "8400020777",
    "9120835960",
    "7024755475",
    "9654461460",
    "7056838487",
    "7295013977",
    "9044947032",
    "9229828718",
    "8755036298",
    "9516884397",
    "9528182175",
    "8369903711",
    "7319176066",
    "9899187565",
    "8077607951",
    "9306493535",
    "9528897029",
    "7737275442",
    "7044767364",
    "9027128203",
    "9311342918",
    "9998950075",
    "9562338704",
    "7835074390",
    "9310187848",
    "6003212022",
    "8860654518",
    "8076060645",
    "7819840393",
    "8265945411",
    "8851336142",
    "9468914587",
    "8810020648",
    "8796942567",
    "8604566463",
    "7550958516",
    "8979958734",
    "7453984273",
    "8695116154",
    "7617599812",
    "9991592274",
    "8080567449",
    "9311576913",
    "7895610761",
    "7988405123",
    "9523584620",
    "7350995377",
    "7042159021",
    "9110015249",
    "7982921084",
    "8791482692",
    "9079206056",
    "9811694731",
    "9739723461",
    "9302163517",
    "9065377666",
    "9352032663",
    "8684869202",
    "7206788849",
    "8146293309",
    "9343048086",
    "7895416920",
    "8168762554",
    "7248846941",
    "7079529736",
    "7701883225",
    "7905704507",
    "8127532736",
    "9235994248",
    "6396761329",
    "9030337409",
    "6235351838",
    "9671606849",
    "8434962569",
    "9811264547",
    "7983822872",
    "9555591644",
    "8750108600",
    "8941949653",
    "8505903177",
    "9830305008",
    "8941050183",
    "7290994428",
    "7428449271",
    "7053630229",
    "8527835099",
    "9996864295",
    "9717832950",
    "8860240669",
    "9351286098",
    "7505702483",
    "9956227007",
    "9179334450",
    "9319063370",
    "8789708201",
    "8742987558",
    "8448834430",
    "8077605026",
    "8586065547",
    "9871569071",
    "9315017768",
    "7906271432",
    "7007703793",
    "7497836503",
    "8806022691",
    "8052907423",
    "7838600857",
    "8595557275",
    "9810943699",
    "7488431657",
    "7008713115",
    "9717593314",
    "8882171320",
    "8233688282",
    "6299915326",
    "9552253129",
    "9718407874",
    "8130062256",
    "8800915623",
    "9871197556",
    "8287845178",
    "8480823037",
    "7840076786",
    "9582818839",
    "9389455793",
    "7324057919",
    "8882195668",
    "6399079303",
    "7289836786",
    "9625063201",
    "9162808516",
    "7557732061",
    "8808590621",
    "9953355111",
    "8920287982",
    "9891545157",
    "9873411235",
    "8810552279",
    "7209799377",
    "7838320333",
    "7240173177",
    "8171195605",
    "8882209913",
    "9654670017",
    "8177044354",
    "7988584465",
    "9729078185",
    "8630029481",
    "7037016941",
    "8920799160",
    "8368594197",
    "8750962795",
    "8930070918",
    "8882546073",
    "9286269684",
    "8619447048",
    "7973801605",
    "9910543884",
    "9720991611",
    "8826709053",
    "7217324752",
    "8709725365",
    "8218429658",
    "7669513632",
    "9315729432",
    "9818307426",
    "7701877704",
    "9599899932",
    "9217065114",
    "9762216220",
    "8847073668",
    "8059928678",
    "6201314365",
    "9910238414",
    "9718444834",
    "8171485680",
    "7700015254",
    "8929964621",
    "7827591060",
    "7538996191",
    "9053793414",
    "8955298317",
    "7817805104",
    "9634357118",
    "6398061843",
    "7982734063",
    "9675563676",
    "7895376326",
    "8221820247",
    "9218057262",
    "7409626882",
    "9129806774",
    "7250314762",
    "9334129100",
    "8655383978",
    "6204822586",
    "9760435492",
    "8757305444",
    "7428657925",
    "8700384618",
    "8084959137",
    "9250288863",
    "8802032090",
    "9711241007",
    "8130287050",
    "8679627946",
    "9756703207",
    "9682070751",
    "7065309919",
    "6201279308",
    "9315673500",
    "8791330303",
    "6205346480",
    "7055550856",
    "9717104363",
    "7091365101",
    "9625238518",
    "9599624049",
    "7027329164",
    "9855924442",
    "8607196833",
    "7903129790",
    "9999303499",
    "8527688640",
    "7078280186",
    "9877160538",
    "6239786068",
    "7011282262",
    "9643743162",
    "7061798230",
    "8851242244",
    "6398360412",
    "8476813238",
    "7678438073",
    "7895206279",
    "8920402326",
    "6370833990",
    "9643081121",
    "9671531408",
    "9855294153",
    "9001720745",
    "9813607530",
    "8130286882",
    "7297038964",
    "7903458829",
    "8882940265",
    "7524803766",
    "7217753019",
    "9627712298",
    "7060796369",
    "6207125837",
    "8800802201",
    "7877962567",
    "9753505912",
    "9555331809",
    "8668747488",
    "7619983792",
    "7367917029",
    "9310134262",
    "7715862882",
    "8684919484",
    "8059579188",
    "8377984206",
    "9412517126",
    "9458273115",
    "8756597906",
    "8376849888",
    "6003103483",
    "9304093474",
    "9911237523",
    "9263737404",
    "9693755131",
    "7011616996",
    "9716972397",
    "8453222643",
    "8431502280",
    "9754424802",
    "9284233119",
    "9434477688",
    "7022136240",
    "9653163602",
    "8076138707",
    "8700655327",
    "9958531230",
    "9560949664",
    "7071307749",
    "8929548768",
    "7906613162",
    "9650557241",
    "8800934517",
    "8888990587",
    "7206200183",
    "9323951769",
    "9929753091",
    "8800805472",
    "8295590991",
    "8532876605",
    "9140453371",
    "7698198073",
    "9899917576",
    "7017488581",
    "6201424737",
    "7905390567",
    "9354176014",
    "8948260990",
    "9068031305",
    "7268967293",
    "8860336652",
    "7779966737",
    "8586048618",
    "7970331142",
    "8178460377",
    "8800616530",
    "9770975031",
    "7428614797",
    "8882879815",
    "6291082159",
    "8178457368",
    "9650970665",
    "7379542718",
    "9887692354",
    "6205231749",
    "7073838373",
    "9873385738",
    "9999507282",
    "7701897107",
    "9873379700",
    "7665283602",
    "9711352713",
    "8851869298",
    "9315978226",
    "9149304893",
    "7827584675",
    "8447045265",
    "7982338398",
    "9560914890",
    "9811898626",
    "8383929182",
    "9711161387",
    "8239142359",
    "8434201280",
    "6307360636",
    "9939955388",
    "9352050712",
    "9151311821",
    "8852835311",
    "8979485024",
    "7827246145",
    "8456842432",
    "8006707973",
    "9800693765",
    "8002444188",
    "9772033690",
    "9199558823",
    "8340183449",
    "9210587650",
    "9368870335",
    "9516462017",
    "9661178494",
    "9601770037",
    "7983067478",
    "9773133226",
    "9899699822",
    "7564083668",
    "9643856461",
    "8210598592",
    "7017608108",
    "8250955115",
    "8826612470",
    "8252313866",
    "8789923242",
    "9053344673",
    "7906855687",
    "9096973029",
    "8318424342",
    "7417347313",
    "8360829252",
    "8802415278",
    "8948260178",
    "7206991718",
    "9654905793",
    "9893359318",
    "7651879083",
    "7847984502",
    "9717421810",
    "7341176755",
    "8218701270",
    "9625597374",
    "8949059450",
    "7668049281",
    "9557144260",
    "9559311624",
    "9310037054",
    "9627673430",
    "7895202588",
    "9090909900",
    "8884983492",
    "8448675703",
    "7979899599",
    "8218696913",
    "8448552769",
    "7725861584",
    "7055040955",
    "7017890403",
    "8755728295",
    "8448976946",
    "7251006297",
    "9097588132",
    "9811512368",
    "6203083640",
    "7992356918",
    "7060558642",
    "8698835643",
    "9915282433",
    "6266233664",
    "8652764081",
    "9560438348",
    "9829314678",
    "7999488848",
    "9350025958",
    "9952013147",
    "9131412144",
    "8271228346",
    "7027827777",
    "9166570165",
    "9999238158",
    "8195832826",
    "9910347108",
    "9784370057",
    "8959710980",
    "9618217660",
    "8130816501",
    "9415504503",
    "9131698337",
    "8957210472",
    "7982696085",
    "8447454923",
    "9050700091",
    "9135668291",
    "6260989815",
    "8882061761",
    "8076499641",
    "7417417811",
    "6396739826",
    "8078633799",
    "8595886842",
    "8603425091",
    "8396830183",
    "8979276983",
    "9887600177",
    "9306940209",
    "8468087097",
    "9665893456",
    "9996610278",
    "8791567288",
    "7988467848",
    "9911688671",
    "9311042021",
    "8527795640",
    "6206294714",
    "9038517513",
    "9716061406",
    "9631709392",
    "8340414612",
    "8207432665",
    "7505286773",
    "7970394200",
    "6201602854",
    "6203588059",
    "9982702045",
    "9142973909",
    "9344822190",
    "6397279790",
    "9052598401",
    "9891778677",
    "9660219898",
    "7480875427",
    "9036730509",
    "6239078892",
    "7065828655",
    "9671949703",
    "6302669359",
    "9996772787",
    "9211724080",
    "8882952514",
    "9365931127",
    "7701909945",
    "9119782202",
    "9639027207",
    "8941033169",
    "8860903858",
    "9369040966",
    "9996110602",
    "9711910394",
    "8800386786",
    "8595479718",
    "9205284737",
    "9582181598",
    "8059095166",
    "9068877684",
    "7428107218",
    "8770575093",
    "9235823247",
    "7382654537",
    "7009348561",
    "7065676997",
    "6370528004",
    "8375945126",
    "7838524954",
    "8595755889",
    "9582393772",
    "8700737672",
    "7014003720",
    "7798032646",
    "9315430594",
    "9711377106",
    "9045678420",
    "9716684129",
    "9634170423",
    "8085467109",
    "8319044435",
    "8851505658",
    "6396255951",
    "9985099672",
    "7249707970",
    "9351922966",
    "8860033237",
    "8651842878",
    "9702441796",
    "9801206040",
    "8882944910",
    "9149007932",
    "8199924850",
    "8053620654",
    "8310965996",
    "7317541262",
    "7310515883",
    "9899137513",
    "9717181681",
    "7204730206",
    "9808962973",
    "9306966545",
    "7088328440",
    "6398271912",
    "7079732600",
    "9004704130",
    "8287995509",
    "8210904327",
    "8700305536",
    "9395578853",
    "8008962718",
    "9541000111",
    "9711693061",
    "7018965940",
    "9217029235",
    "8802257930",
    "8860880333",
    "8755568987",
    "7078714513",
    "9759589403",
    "7494941131",
    "9576622043",
    "8604444217",
    "8079093226",
    "8607319115",
    "9654203620",
    "8527155809",
    "8638003453",
    "9870105511",
    "7417405306",
    "7456977698",
    "8396906712",
    "9817082516",
    "8789105948",
    "8447024679",
    "9810550930",
    "7903853571",
    "8171854044",
    "6395694104",
    "8750901382",
    "8299481397",
    "9315354930",
    "8077363036",
    "8791780006",
    "7988704950",
    "9650316762",
    "8826505129",
    "9559470400",
    "8085139004",
    "6239701944",
    "8527062691",
    "7376788089",
    "8059954062",
    "9355056969",
    "9045245702",
    "7668184858",
    "7027698288",
    "8178932976",
    "7248495968",
    "7065672006",
    "7015188649",
    "8294248080",
    "9125538715",
    "9305481510",
    "8077758311",
    "9955079812",
    "9899881514",
    "9131979939",
    "8368016496",
    "7905425221",
    "9508538994",
    "9711877683",
    "7999302535",
    "9082466956",
    "7042407306",
    "6206887949",
    "8468823273",
    "6396928568",
    "9958872749",
    "9971476059",
    "9289359665",
    "6002130724",
    "9518856356",
    "8920438086",
    "8872830553",
    "9608956718",
    "6388648183",
    "8218003058",
    "9997366068",
    "9958589192",
    "8397973997",
    "9729201292",
    "9319897193",
    "9549860629",
    "9588715065",
    "9568927796",
    "8168320955",
    "8171660245",
    "8130302360",
    "9306974269",
    "6376202536",
    "7300270911",
    "9761669941",
    "8448424692",
    "9053379902",
    "8708814095",
    "8887596198",
    "7759821454",
    "7838957437",
    "9354075791",
    "9773670744",
    "9354618330",
    "9467417418",
    "6206538016",
    "9518260282",
    "8540088798",
    "8750814119",
    "9625724065",
    "8057699679",
    "8630670512",
    "8000540184",
    "9118319707",
    "8826901877",
    "7004886481",
    "6396528522",
    "9565328892",
    "9217901330",
    "9693464384",
    "9873517878",
    "8076986001",
    "8685818284",
    "9756434481",
    "7042895508",
    "8210243796",
    "8527897509",
    "7303559036",
    "7081433393",
    "7599660205",
    "6376545272",
    "9235938204",
    "8860942577",
    "9027948849",
    "8799237191"
    ]



class AffliateExcelUpdateView(APIView):
    def post(self, request, format=None):
        print(len(phone_key))
        dd = (
            DossierData.objects
            .filter(phone__in=phone_key, source=15)
            .order_by('phone', '-id')
            .distinct('phone').values("id","phone","source","utm_source","utm_medium")
        )
        print(dd)
        print()
        # for i in dd:
        #     print(i["id"])

        sheet = get_google_sheet_aeutplp()
        i = 0
        for lobj in dd: 
            try:
                selected_date = "aabc"
                row_data = [
                    lobj["utm_source"],
                    lobj["utm_medium"]
                ]

                # find email in column B
                cell = sheet.find(lobj["phone"])

                if cell:
                    row_number = cell.row
                    sheet.update(f"A{row_number}:B{row_number}", [row_data])
                    print(f"Row {row_number} updated successfully")

                    print("row updated successfully")

                else:
                    print("row not found, new row inserted")
            except Exception as e:
                print("google sheet error", str(e))
            i+=1
            print(i)
        return Response({"msg":"User Created Successfully"})
    


