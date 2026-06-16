from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import *
from .serializers import *
from rest_framework import filters
from gcc_backend.pagination import CustomPageNumberPagination
from rest_framework.permissions import IsAuthenticated
from django.conf import settings
from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template
from google.cloud import storage
import os
from gcc_backend.utils import *
import pandas as pd
import tempfile
import re
from datetime import datetime, timedelta
from django.utils.dateparse import parse_date
client = storage.Client(project=settings.GS_PROJECT_ID)
import io
from openpyxl import Workbook
from django.core.mail import EmailMessage
from django.utils.timezone import now
from rest_framework import status
from django.utils import timezone
from datetime import timedelta
from django.db.models import F, FloatField, ExpressionWrapper
from django.db.models.functions import Cast

# Create your views here.

class StudentQuery_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id']
    ordering_fields = ['id']
    def get(self, request):
        datas = StudentEnquiries.objects.all().order_by('-id')

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListStudentQuerySerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    


class StudentData_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id']
    ordering_fields = ['id']
    def get(self, request):
        datas = StudentsData.objects.all().order_by('-id')

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListStudentDataSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    


class StudentPayment_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"razorpay_order_id","razorpay_payment_id","amount","dossier_form__full_name","dossier_form__email","dossier_form__phone","dossier_form__state","dossier_form__city","status"]
    ordering_fields = ['id',"created_at","dossier_form__full_name","dossier_form__email","dossier_form__phone","dossier_form__state","dossier_form__city","razorpay_order_id","razorpay_payment_id","amount"]
    def get(self, request):

        datas = Payments.objects.filter(source=SourceType.Website).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(dossier_form__full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(dossier_form__email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(dossier_form__phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(dossier_form__state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(dossier_form__city__icontains=city)

        university = request.GET.get('university')
        if university:
            datas = datas.filter(dossier_form__university__icontains=university)
        fee_waiver_category = request.GET.get('fee_waiver_category')
        if fee_waiver_category:
            datas = datas.filter(fee_waiver_category__icontains=fee_waiver_category)

        status = request.GET.get('status')
        if status:
            datas = datas.filter(status__icontains=status)

        
        # Date range filter
        start_date = request.GET.get('start_date')
        end_date   = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListStudentPaymentSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    

class StudentSourcePayment_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id',"razorpay_order_id","razorpay_payment_id","amount","dossier_form__full_name","dossier_form__email","dossier_form__phone","dossier_form__state","dossier_form__city","status"]
    ordering_fields = ['id',"created_at","dossier_form__full_name","dossier_form__email","dossier_form__phone","dossier_form__state","dossier_form__city","razorpay_order_id","razorpay_payment_id","amount"]
    def get(self, request):
        source_type = request.GET.get("source")
        if source_type:
            datas = Payments.objects.filter(source=source_type).order_by('-id')
        else:
            datas = Payments.objects.filter(source=SourceType.Website).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(dossier_form__full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(dossier_form__email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(dossier_form__phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(dossier_form__state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(dossier_form__city__icontains=city)

        status = request.GET.get('status')
        if status:
            datas = datas.filter(status__icontains=status)

        university = request.GET.get('university')
        if university:
            datas = datas.filter(dossier_form__university__icontains=university)
        fee_waiver_category = request.GET.get('fee_waiver_category')
        if fee_waiver_category:
            datas = datas.filter(fee_waiver_category__icontains=fee_waiver_category)

            
        # Date range filter
        start_date = request.GET.get('start_date')
        end_date   = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListStudentPaymentSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    




class ExportPaymentExcelView(APIView):
    def get(self, request):
        recipient_email = "testtechno0@yopmail.com"
        # 1. Fetch data (Reuse your annotated logic)
        # Note: Using .iterator() is better for large datasets to save memory
        
        # 1. Calculate the timestamp for 24 hours ago
        time_threshold = timezone.now() - timedelta(hours=24)

        # 2. Filter queryset
        # Note: 'isnull' is lowercase in Django lookups
        queryset = Payments.objects.filter(
            razorpay_order_id__isnull=False,
            created_at__gte=time_threshold
        ).order_by('-created_at')
        # 2. Create Excel in memory
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments Report"

        # Headers
        headers = ['Order ID','Payment ID', 'Amount', 'Currency', 'Student Name', 'Email', "Phone","City","State","Payment Status",'Date']
        ws.append(headers)

        # Rows
        for p in queryset:

            dossier = p.dossier_form

            ws.append([
                p.razorpay_order_id,
                p.razorpay_payment_id,
                float(p.amount),
                p.currency,
                dossier.full_name if dossier else "N/A",
                dossier.email if dossier else "N/A",
                dossier.phone if dossier else "N/A",
                dossier.city if dossier else "N/A",
                dossier.state if dossier else "N/A",
                p.status,
                p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else "N/A"
            ])

        # Save to BytesIO stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)


        queryset1 = DossierData.objects.filter(
            created_at__gte=time_threshold
        ).order_by('-created_at')
        # 2. Create Excel in memory
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Dossior Lead Report"

        # Headers
        headers = ['Student Name', 'Email', "Phone","City","State",'Date']
        ws1.append(headers)

        # Rows
        for dossier in queryset1:

            ws1.append([
                dossier.full_name if dossier else "N/A",
                dossier.email if dossier else "N/A",
                dossier.phone if dossier else "N/A",
                dossier.city if dossier else "N/A",
                dossier.state if dossier else "N/A",
                dossier.created_at.strftime('%Y-%m-%d %H:%M') if dossier.created_at else "N/A"
            ])

        # Save to BytesIO stream
        output1 = io.BytesIO()
        wb1.save(output1)
        output1.seek(0)


        queryset2 = ContactUs.objects.filter(
            created_at__gte=time_threshold
        ).order_by('-created_at')
        # 2. Create Excel in memory
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Contact Us Report"

        # Headers
        headers = ['First Name','Last Name', 'Email', "Phone","City","State",'Date']
        ws2.append(headers)

        # Rows
        for dossier in queryset2:

            ws2.append([
                dossier.first_name if dossier else "N/A",
                dossier.last_name if dossier else "N/A",
                dossier.email if dossier else "N/A",
                dossier.phone if dossier else "N/A",
                dossier.city if dossier else "N/A",
                dossier.state if dossier else "N/A",
                dossier.created_at.strftime('%Y-%m-%d %H:%M') if dossier.created_at else "N/A"
            ])

        # Save to BytesIO stream
        output2 = io.BytesIO()
        wb2.save(output2)
        output2.seek(0)

        # 3. Send Email
        try:
            subject = f"Payment Report, Dossier Lead & Contact Us - {now().strftime('%d %b %Y')}"

            bcc_list = ['atul.tevatia@kcglobed.com',"harish.kumar@kcglobed.com"]
            # Corrected the slashes to backslashes for proper line breaks
            message = (
                "Hello Sir,\n\n"
                "Please find the attached payment report , dossier lead report & Contact Us for the last 24 hours.\n\n"
                "Thanks,\n"
                "KCGlobed Team"
            )
            email = EmailMessage(
                subject,
                message,
                'info@gccschool.com',
                ["info@gccschoo.com","nfet@gccschool.com"],
                bcc=bcc_list,
            )
            
            # Attach the file (filename, content, mimetype)
            email.attach(
                f'Payment_Report_{now().strftime("%Y%m%d")}.xlsx',
                output.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            email.attach(
                f'Dossier_Report_{now().strftime("%Y%m%d")}.xlsx',
                output1.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            email.attach(
                f'Contact_Us_Report_{now().strftime("%Y%m%d")}.xlsx',
                output2.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            email.send()

            return Response({"message": "Email sent successfully"}, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CampusFaculty_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['full_name',"email","mobile","state","city"]
    ordering_fields = ["id",'full_name',"email","mobile","state","city","created_at"]
    def get(self, request):
        datas = CampusFaculty.objects.all().order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        mobile = request.GET.get('mobile')
        if mobile:
            datas = datas.filter(mobile__icontains=mobile)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        
        address = request.GET.get('address')
        if address:
            datas = datas.filter(address__icontains=address)

        institution_name = request.GET.get('institution_name')
        if institution_name:
            datas = datas.filter(institution_name__icontains=institution_name)


        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListCampusFacultySerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)


class GetFacultyCampusReportPDFView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['full_name',"email","mobile","state","city"]
    ordering_fields = ["id",'full_name',"email","mobile","state","city","created_at"]
    def get(self, request):
        datas = CampusFaculty.objects.all().order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        mobile = request.GET.get('mobile')
        if mobile:
            datas = datas.filter(mobile__icontains=mobile)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        address = request.GET.get('address')
        if address:
            datas = datas.filter(address__icontains=address)

        institution_name = request.GET.get('institution_name')
        if institution_name:
            datas = datas.filter(institution_name__icontains=institution_name)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListCampusFacultySerializer(datas, many=True)

        data = {
                    "user_data":serializers.data,
                    "report_date": datetime.now().strftime("%d-%m-%Y, %H:%M")
                }
        

        template = get_template('pdf/faculty_campus_report.html')
        html  = template.render(data)
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Encode HTML and create PDF
            html = html.encode('latin-1', 'replace').decode('latin-1')
            pdf = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=temp_file)

            if pdf.err:
                raise Exception("PDF generation error!")
        
        # After the 'with' block, the file is closed, but not deleted yet
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "faculty_campus_report"
            gcs_folder_name = "media/gcc_reports"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.pdf"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted from the server's disk
            os.remove(pdf_path)
    


class GetFacultyCampusReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['full_name',"email","mobile","state","city"]
    ordering_fields = ["id",'full_name',"email","mobile","state","city","created_at"]
    def get(self, request):
        datas = CampusFaculty.objects.all().order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        mobile = request.GET.get('mobile')
        if mobile:
            datas = datas.filter(mobile__icontains=mobile)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        address = request.GET.get('address')
        if address:
            datas = datas.filter(address__icontains=address)

        institution_name = request.GET.get('institution_name')
        if institution_name:
            datas = datas.filter(institution_name__icontains=institution_name)
               
        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListCampusFacultySerializer(datas, many=True)

        lis = []
        
        lis.append({
                "name":"Faculty Campus Report",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "total_questions":'',
                "institution_name":"",
                "department":"",
                "designation":"",
                "teaching_experience":"",
                "industrial_experience":"",
                "highest_qualification":"",
                "motivation":"",
                "support_activities":"",
                "student_reach":"",
                "created_at":""
            })

       
        lis.append({
                "name":"",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "total_questions":'',
                "institution_name":"",
                "department":"",
                "designation":"",
                "teaching_experience":"",
                "industrial_experience":"",
                "highest_qualification":"",
                "motivation":"",
                "support_activities":"",
                "student_reach":"",
                "created_at":""
            })
        
        lis.append({
                "name":"Full Name",
                "email":'Email',
                "subject":'Phone Number',
                "Chapter":'City',
                "Topic":'State',
                "total_questions":'Address',
                "institution_name":"Institution Name",
                "department":"Department",
                "designation":"Designation",
                "teaching_experience":"Teaching Experience",
                "industrial_experience":"Industrial Experience",
                "highest_qualification":"Highest Qualification",
                "motivation":"Motivation",
                "support_activities":"Support Activities",
                "student_reach":"Student Reach",
                "created_at":"Created At"
            })
        
        
        for chapter_data in serializers.data:

            activities = chapter_data['support_activities']
            activities_str = ", ".join(map(str, activities)) if isinstance(activities, list) else str(activities)


            lis.append({
                "name":chapter_data['full_name'],
                "email":chapter_data['email'],
                "subject":chapter_data['mobile'],
                "Chapter":chapter_data['city'],
                "Topic":chapter_data['state'],
                "total_questions":chapter_data['address'],
                "institution_name":chapter_data['institution_name'],
                "department":chapter_data['department'],
                "designation":chapter_data['designation'],
                "teaching_experience":chapter_data['teaching_experience'],
                "industrial_experience":chapter_data['industrial_experience'],
                "highest_qualification":chapter_data['highest_qualification'],
                "motivation":chapter_data['motivation'],
                "support_activities":activities_str,
                "student_reach":chapter_data['student_reach'],
                "created_at":chapter_data['created_at']
            })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "faculty_campus_report"
            gcs_folder_name = "media/gcc_reports"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)


class CampusStudent_list(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['full_name',"email","mobile","state","city"]
    ordering_fields = ["id",'full_name',"email","mobile","state","city","created_at"]
    def get(self, request):
        datas = CampusStudent.objects.all().order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        mobile = request.GET.get('mobile')
        if mobile:
            datas = datas.filter(mobile__icontains=mobile)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        address = request.GET.get('address')
        if address:
            datas = datas.filter(address__icontains=address)

        college_name = request.GET.get('college_name')
        if college_name:
            datas = datas.filter(college_name__icontains=college_name)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ListCampusStudentSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    


class GetStudentCampusReportPDFView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['full_name',"email","mobile","state","city"]
    ordering_fields = ["id",'full_name',"email","mobile","state","city","created_at"]
    def get(self, request):
        datas = CampusStudent.objects.all().order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        mobile = request.GET.get('mobile')
        if mobile:
            datas = datas.filter(mobile__icontains=mobile)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        address = request.GET.get('address')
        if address:
            datas = datas.filter(address__icontains=address)

        college_name = request.GET.get('college_name')
        if college_name:
            datas = datas.filter(college_name__icontains=college_name)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListCampusStudentSerializer(datas, many=True)

        data = {
                    "user_data":serializers.data,
                    "report_date": datetime.now().strftime("%d-%m-%Y, %H:%M")
                }
        

        template = get_template('pdf/student_campus_report.html')
        html  = template.render(data)
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Encode HTML and create PDF
            html = html.encode('latin-1', 'replace').decode('latin-1')
            pdf = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=temp_file)

            if pdf.err:
                raise Exception("PDF generation error!")
        
        # After the 'with' block, the file is closed, but not deleted yet
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "student_campus_report"
            gcs_folder_name = "media/gcc_reports"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.pdf"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted from the server's disk
            os.remove(pdf_path)
    


class GetStudentCampusReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['full_name',"email","mobile","state","city"]
    ordering_fields = ["id",'full_name',"email","mobile","state","city","created_at"]
    def get(self, request):
        datas = CampusStudent.objects.all().order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        mobile = request.GET.get('mobile')
        if mobile:
            datas = datas.filter(mobile__icontains=mobile)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        address = request.GET.get('address')
        if address:
            datas = datas.filter(address__icontains=address)

        college_name = request.GET.get('college_name')
        if college_name:
            datas = datas.filter(college_name__icontains=college_name)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ListCampusStudentSerializer(datas, many=True)

        lis = []
        
        lis.append({
                "name":"Student Campus Report",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "total_questions":'',
                "institution_name":"",
                "department":"",
                "designation":"",
                "teaching_experience":"",
                "industrial_experience":"",
                "highest_qualification":"",
                "motivation":"",
                "support_activities":"",
                "student_reach":"",
                "created_at":""
            })

       
        lis.append({
                "name":"",
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "total_questions":'',
                "institution_name":"",
                "department":"",
                "designation":"",
                "teaching_experience":"",
                "industrial_experience":"",
                "highest_qualification":"",
                "motivation":"",
                "support_activities":"",
                "student_reach":"",
                "created_at":""
            })
        
        lis.append({
                "name":"Full Name",
                "email":'Email',
                "subject":'Phone Number',
                "Chapter":'City',
                "Topic":'State',
                "total_questions":'Address',
                "institution_name":"College Name",
                "department":"Study Program",
                "designation":"Study Program Other",
                "teaching_experience":"Semester",
                "industrial_experience":"Student Body Member",
                "highest_qualification":"Campus Ambassador History",
                "motivation":"Inspiration",
                "support_activities":"Promotion Channels",
                "student_reach":"Student Reach",
                "created_at":"Created At"
            })
        
        
        for chapter_data in serializers.data:
            activities = chapter_data['promotion_channels']
            activities_str = ", ".join(map(str, activities)) if isinstance(activities, list) else str(activities)

            lis.append({
                "name":chapter_data['full_name'],
                "email":chapter_data['email'],
                "subject":chapter_data['mobile'],
                "Chapter":chapter_data['city'],
                "Topic":chapter_data['state'],
                "total_questions":chapter_data['address'],
                "institution_name":chapter_data['college_name'],
                "department":chapter_data['program_of_study'],
                "designation":chapter_data['program_other'],
                "teaching_experience":chapter_data['semester'],
                "industrial_experience":chapter_data['student_body_member'],
                "highest_qualification":chapter_data['campus_ambassador_history'],
                "motivation":chapter_data['inspiration'],
                "support_activities":activities_str,
                "student_reach":chapter_data['student_reach'],
                "created_at":chapter_data['created_at']
            })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "student_campus_report"
            gcs_folder_name = "media/gcc_reports"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)



class ContactUsView(APIView):
    def post(self, request, format=None):
        serializer = ContactUsSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            user  = serializer.save()
            return Response({'message':'Message sent Successfully','data':serializer.data})

        return Response(serializer.errors)
    

class GetContactUSView(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name',"last_name","email","phone","state","city"]
    ordering_fields = ["id",'first_name',"last_name","email","phone","state","city","created_at"]
    def get(self, request):
        datas = ContactUs.objects.all().order_by("-id")

        first_name = request.GET.get('first_name')
        if first_name:
            datas = datas.filter(first_name__icontains=first_name)

        last_name = request.GET.get('last_name')
        if last_name:
            datas = datas.filter(last_name__icontains=last_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = ContactListSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    

class GetContactusReportPDFView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name',"last_name","email","phone","state","city"]
    ordering_fields = ["id",'first_name',"last_name","email","phone","state","city","created_at"]
    def get(self, request):
        datas = ContactUs.objects.all().order_by("-id")

        first_name = request.GET.get('first_name')
        if first_name:
            datas = datas.filter(first_name__icontains=first_name)

        last_name = request.GET.get('last_name')
        if last_name:
            datas = datas.filter(last_name__icontains=last_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ContactListSerializer(datas, many=True)


        data = {
                    "user_data":serializers.data,
                    "report_date": datetime.now().strftime("%d-%m-%Y, %H:%M")
                }
        

        template = get_template('pdf/contact_us_report.html')
        html  = template.render(data)
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Encode HTML and create PDF
            html = html.encode('latin-1', 'replace').decode('latin-1')
            pdf = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=temp_file)

            if pdf.err:
                raise Exception("PDF generation error!")
        
        # After the 'with' block, the file is closed, but not deleted yet
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "contact_us_report"
            gcs_folder_name = "media/gcc_reports"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.pdf"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted from the server's disk
            os.remove(pdf_path)
    


class GetContactusReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name',"last_name","email","phone","state","city"]
    ordering_fields = ["id",'first_name',"last_name","email","phone","state","city","created_at"]
    def get(self, request):
        datas = ContactUs.objects.all().order_by("-id")

        first_name = request.GET.get('first_name')
        if first_name:
            datas = datas.filter(first_name__icontains=first_name)

        last_name = request.GET.get('last_name')
        if last_name:
            datas = datas.filter(last_name__icontains=last_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)

        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)
            
        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        serializers = ContactListSerializer(datas, many=True)

        lis = []
        
        lis.append({
                "name":"Contact Us Report",
                "last_name":'',
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "total_questions":''
            })

       
        lis.append({
                "name":"",
                "last_name":'',
                "email":'',
                "subject":'',
                "Chapter":'',
                "Topic":'',
                "total_questions":''
            })
        
        lis.append({
                "name":"First Name",
                "last_name":'Last Name',
                "email":'Email',
                "subject":'Phone Number',
                "Chapter":'City',
                "Topic":'State',
                "total_questions":'Created At',
            })
        
        
        for chapter_data in serializers.data:
            lis.append({
                "name":chapter_data['first_name'],
                "last_name":chapter_data['last_name'],
                "email":chapter_data['email'],
                "subject":chapter_data['phone'],
                "Chapter":chapter_data['city'],
                "Topic":chapter_data['state'],
                "total_questions":chapter_data['created_at'],
            })


        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(lis)
            df.to_excel(pdf_path, header=False, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "contact_us_report"
            gcs_folder_name = "media/gcc_reports"
            gcs_file_name = f"{gcs_folder_name}/{report_name}_{timestamp}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)

            return success_response(
                message="Success",
                data={"report_url": blob.public_url},
                status_code=status.HTTP_200_OK
            )
        finally:
            # Ensure the temporary file is deleted
            os.remove(pdf_path)



class CreateStudentProfileView(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = CompleteStudentSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            user  = serializer.save()
            return Response({'message':'Message sent Successfully','data':[]})
        return Response(serializer.errors)


class CreateStudentProfileDraftView(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        print("calling....")
        print(request.data)
        serializer = CompleteStudentDraftSerializer(data = request.data)
        if serializer.is_valid(raise_exception = True):
            user  = serializer.save()
            return Response({'message':'Message sent Successfully','data':[]})
        return Response(serializer.errors)


class StudentSlotBookView(APIView):
    permission_classes = [IsAuthenticated]
    def patch(self, request):
        datas = StudentProfile.objects.filter(user = request.user).first()
        if datas is not None:
            serializers = StudentSlotBookSerializer(datas, data=request.data, partial=True)
            if serializers.is_valid():
                serializers.save()
                return Response({"message": "Success","status":200, "data": {}})
            return Response({'message':'failed','status':400,'data':serializers.errors})

        return Response({'message':'failed','status':400, 'data':[]})
    

class StudentMockTestCompleteStatusView(APIView):
    def post(self, request):
        datas = StudentProfile.objects.filter(application_id=request.data["email"])
        if not datas:
            return Response({'message':'Invalid Account','status':400,'data':[]})
        datas = datas.first()
        serializers = StudentMockTestCompleteStatusSerializer(datas, data=request.data, partial=True)
        if serializers.is_valid():
            serializers.save()
            return Response({"message": "Success","status":200, "data": []})
        return Response({'message':'failed','status':400, 'data':[serializers.errors]})
    

class StudentApplicationIdUpdateView(APIView):
    def post(self, request):
        usr = User.objects.all()
        for i in usr:
            StudentProfile.objects.filter(user=i).update(application_id=i.application_id)
            # print(std_profile)
        return Response({"message": "Success","status":200, "data": []})
    

class StudentMockTestStartStatusView(APIView):
    def post(self, request):
        datas = StudentProfile.objects.filter(email=request.data["email"])
        if not datas:
            return Response({'message':'Invalid Account','status':400,'data':[]})
        datas = datas.first()
        serializers = StudentMockTestStartStatusSerializer(datas, data=request.data, partial=True)
        if serializers.is_valid():
            serializers.save()
            return Response({"message": "Success","status":200, "data": []})
        return Response({'message':'failed','status':400, 'data':[serializers.errors]})
    

class GetStudentAdmitCardView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        std_data = StudentProfile.objects.filter(user = request.user).first()
        static_selected_bucket = settings.GS_BUCKET_NAME
        context = {
            "username": request.user.email,
            "user_id": request.user.id,
            "application_id": request.user.application_id,
            "student_name": std_data.first_name+" "+std_data.last_name,
            "slot_date": std_data.slot_date,
            "slot_time": std_data.slot_time,
            "photo": std_data.photo.url,
            "signature": std_data.signature.url,
            "barcode":"",
            "report_date": datetime.now(),
            # "test_link":"https://cocubes.in/gccschool-nfet",
            "test_link":"https://www.gccschool.com/myaccount",
            "bucket_static_logo":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/gcc-admit-card-logo.jpeg",
            # "bucket_static_signature":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/admit_card_signature.png"
            "bucket_static_signature":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/gcc_admit_card_sign.png"
        }
        # Render template
        template = get_template("pdf/student_admit_card.html")
        html = template.render(context)

        # xhtml2pdf needs ISO-8859-1
        html = html.encode("ISO-8859-1", "ignore").decode("ISO-8859-1")

        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_path = tmp.name
            pisa_status = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=tmp)

        if pisa_status.err:
            os.remove(pdf_path)
            return Response({"error": "PDF generation failed"}, status=500)
        try:
            # Upload to GCS
            username = re.sub(r"\s+", "_", request.user.email)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            gcs_file = f"media/admit_card/{username}_{request.user.id}.pdf"

            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)
            blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)



class GetStudentAdmitCardAdminView(APIView):
    # permission_classes = [IsAuthenticated]
    def get(self, request, id):
        user_obj = User.objects.filter(id=id).first()
        std_data = StudentProfile.objects.filter(user = user_obj).first()
        static_selected_bucket = settings.GS_BUCKET_NAME
        context = {
            "username": user_obj.email,
            "user_id": user_obj.id,
            "application_id": user_obj.application_id,
            "student_name": std_data.first_name+" "+std_data.last_name,
            "slot_date": std_data.slot_date,
            "slot_time": std_data.slot_time,
            "photo": std_data.photo.url,
            "signature": std_data.signature.url,
            "barcode":"",
            "report_date": datetime.now(),
            # "test_link":"https://cocubes.in/gccschool-nfet",
            "test_link":"https://www.gccschool.com/myaccount",
            "bucket_static_logo":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/gcc-admit-card-logo.jpeg",
            # "bucket_static_signature":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/admit_card_signature.png"
            "bucket_static_signature":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/gcc_admit_card_sign.png"
        }
        # Render template
        template = get_template("pdf/student_admit_card.html")
        html = template.render(context)

        # xhtml2pdf needs ISO-8859-1
        html = html.encode("ISO-8859-1", "ignore").decode("ISO-8859-1")

        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_path = tmp.name
            pisa_status = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=tmp)

        if pisa_status.err:
            os.remove(pdf_path)
            return Response({"error": "PDF generation failed"}, status=500)
        try:
            # Upload to GCS
            username = re.sub(r"\s+", "_", user_obj.email)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            gcs_file = f"media/admit_card/{username}_{user_obj.id}.pdf"

            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)
            blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)




class GetStudentProfileView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        datas = StudentProfile.objects.filter(user = request.user).first()
        if datas is not None:
            serializers = StudentProfileSerializer(datas)
            return Response({'message':'success',"status":200,'data':serializers.data})

        return Response({'message':'failed',"status":400,'data':[]})
    

class GetStudentProfileDraftView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        datas = StudentProfileDraft.objects.filter(user = request.user).first()
        if datas is not None:
            serializers = StudentProfileDraftSerializer(datas)
            return Response({'message':'success',"status":200,'data':serializers.data})

        return Response({'message':'failed',"status":400,'data':[]})
    


class GetStudentReAttemptView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        status = request.data.get("status")

        if not status:
            return Response({
                "message":"Status is required",
                "status":status.HTTP_400_BAD_REQUEST,
                "data":{}
            })
        datas = StudentProfile.objects.filter(user = request.user).first()
        if datas is not None:
            serializers = StudentReAttemptSerializer(datas, data=request.data, partial=True)
            if serializers.is_valid():
                serializers.save()
                return Response({'message':'success',"status":200,'data':{}})
            else:
                return Response({'message':'failed',"status":400,'data':serializers.errors})
        return Response({'message':'failed',"status":400,'data':{}})
    
from django.db.models import OuterRef, Subquery

    
class GetStudentProfileListingView(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name',"last_name","email","phone","state","city","fee_waiver_category"]
    ordering_fields = ['first_name',"last_name","email","phone","state","city","created_at","slot_date"]
    def get(self, request):
        datas = StudentProfile.objects.all().order_by('-id')

        # Date range filter
        
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)
        
        # slot_date range filter
        start_slot_date = request.GET.get('start_slot_date')
        end_slot_date = request.GET.get('end_slot_date')
        if start_slot_date:
            datas = datas.filter(slot_date__gte=start_slot_date)

        if end_slot_date:
            datas = datas.filter(slot_date__lte=end_slot_date)

        fee_waiver_category = request.GET.get('fee_waiver_category')
        if fee_waiver_category:
            datas = datas.filter(fee_waiver_category=fee_waiver_category)

        first_name = request.GET.get('first_name')
        if first_name:
            datas = datas.filter(first_name__icontains=first_name)
        last_name = request.GET.get('last_name')
        if last_name:
            datas = datas.filter(last_name__icontains=last_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)
        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)
        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)
        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)

        is_result = request.GET.get('is_result')

        # print(is_result)
        if is_result:
            # print(is_result)
            if str(is_result) == "50":
                # latest_results = (
                #     StudentRealExamResult.objects
                #     .order_by('student_profile_id', '-id')
                #     .distinct('student_profile_id')
                #     .annotate(
                #         total_score_float=Cast('totalscore', FloatField()),
                #         total_questions_float=Cast('totalquestions', FloatField()),
                #     )
                #     .annotate(
                #         percentage=ExpressionWrapper(
                #             (F('total_score_float') * 100.0) / F('total_questions_float'),
                #             output_field=FloatField()
                #         )
                #     )
                #     .filter(
                #         percentage__gt=50
                #     )
                #     .values_list(
                #         "student_profile_id",
                #         flat=True
                #     )
                # )
                # datas = datas.filter(id__in=latest_results)

                latest_result_subquery = StudentRealExamResult.objects.filter(
                            student_profile=OuterRef('student_profile')
                        ).order_by('-id')

                latest_results = (
                    StudentRealExamResult.objects
                    .filter(
                        id=Subquery(latest_result_subquery.values('id')[:1])
                    )
                    .annotate(
                        total_score_float=Cast('totalscore', FloatField()),
                        total_questions_float=Cast('totalquestions', FloatField()),
                    )
                    .annotate(
                        percentage=ExpressionWrapper(
                            (F('total_score_float') * 100.0) / F('total_questions_float'),
                            output_field=FloatField()
                        )
                    )
                    .filter(
                        percentage__gt=50
                    )
                    .values_list(
                        'student_profile_id',
                        flat=True
                    )
                )
                datas = datas.filter(id__in=latest_results)

                # return Response({"abc":"abc","data":list(latest_results)})
            
                # datas = datas.filter(
                #         id__in=StudentRealExamResult.objects.annotate(
                #             total_score_float=Cast('totalscore', FloatField()),
                #             total_questions_float=Cast('totalquestions', FloatField()),
                #         ).annotate(
                #             percentage=ExpressionWrapper(
                #                 (F('total_score_float') * 100.0) / F('total_questions_float'),
                #                 output_field=FloatField()
                #             )
                #         ).filter(
                #             percentage__gt=50
                #         ).values_list(
                #             "student_profile_id",
                #             flat=True
                #         )
                #     )

            else:
                # print("datass")
                datas = datas.filter(
                    id__in=StudentRealExamResult.objects.values_list(
                        "student_profile_id",
                        flat=True
                    )
                )

        search_filter = filters.SearchFilter()
        datas = search_filter.filter_queryset(request, datas, self)

        ordering_filter = filters.OrderingFilter()
        datas = ordering_filter.filter_queryset(request, datas, self)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(datas, request, view=self)
        serializers = StudentProfileListSerializer(page, many=True)
        
        return paginator.get_paginated_response(serializers.data)
    




class GetStudentScoreCardView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user_obj = request.user

        std_data = StudentProfile.objects.filter(user=user_obj).first()
        score_objs = StudentRealExamResult.objects.filter(student_profile=std_data).last()

        if not score_objs:
            return Response({"message": "No data found"},data={}, status=404)
        
        datas = []
        for i in score_objs.json_data:
            obj = {}
            obj["Name"] = i["Name"]
            obj["TotalQuestions"] = i["TotalQuestions"]
            obj["Incorrect"] = int(float(i["Attempted"])-float(i['Correct']))
            obj["Correct"] = i["Correct"]
            obj["NotAttempted"] = int(float(i["TotalQuestions"]) - float(i["Attempted"]))
            datas.append(obj)

        static_selected_bucket = settings.GS_BUCKET_NAME
        context = {
            "candidate_name":f"{std_data.first_name.upper()} {std_data.last_name.upper()}",
            "application_id":user_obj.application_id,
            "date_of_exam":std_data.slot_date,
            "time_of_exam":std_data.slot_time,
            "sections":datas,
            "total_questions":score_objs.totalquestions,
            "total_correct":score_objs.totalcorrectanswers,
            "total_incorrect":int(float(score_objs.totalquestionsattempted) - float(score_objs.totalcorrectanswers)),
            "total_not_attempted":int(float(score_objs.totalquestions) - float(score_objs.totalquestionsattempted)),

            "username": user_obj.email,
            "user_id": user_obj.id,
            "application_id": user_obj.application_id,
            "student_name": f'''{std_data.first_name.upper()}" "{std_data.last_name.upper()}''',
            "slot_date": std_data.slot_date,
            "slot_time": std_data.slot_time,
            "photo": std_data.photo.url,
            "signature": std_data.signature.url,
            "barcode":"",
            "report_date": datetime.now(),
            # "test_link":"https://cocubes.in/gccschool-nfet",
            "test_link":"https://www.gccschool.com/myaccount",
            "bucket_static_logo":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/gcc-admit-card-logo.jpeg",
            # "bucket_static_signature":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/admit_card_signature.png"
            "bucket_static_signature":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/gcc_admit_card_sign.png"
        }
        
        # Render template
        template = get_template("pdf/student_score_card.html")
        html = template.render(context)

        # xhtml2pdf needs ISO-8859-1
        html = html.encode("ISO-8859-1", "ignore").decode("ISO-8859-1")

        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_path = tmp.name
            pisa_status = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=tmp)

        if pisa_status.err:
            os.remove(pdf_path)
            return Response({"error": "PDF generation failed"}, status=500)
        try:
            # Upload to GCS
            username = re.sub(r"\s+", "_", user_obj.email)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            gcs_file = f"media/admit_card/{username}_{user_obj.id}.pdf"

            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)
            blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)


class GetAdminStudentScoreCardView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, stid):
        
        std_data = StudentProfile.objects.filter(id=stid).first()
        score_objs = StudentRealExamResult.objects.filter(student_profile=std_data).last()

        if not score_objs:
            return Response({"message": "No data found", "status":404, "data":{}})
        
        datas = []
        for i in score_objs.json_data:
            obj = {}
            obj["Name"] = i["Name"]
            obj["TotalQuestions"] = i["TotalQuestions"]
            obj["Incorrect"] = int(float(i["Attempted"])-float(i['Correct']))
            obj["Correct"] = i["Correct"]
            obj["NotAttempted"] = int(float(i["TotalQuestions"]) - float(i["Attempted"]))
            datas.append(obj)

        static_selected_bucket = settings.GS_BUCKET_NAME
        context = {
            "candidate_name":f"{std_data.first_name.upper()} {std_data.last_name.upper()}",
            "application_id":std_data.application_id,
            "date_of_exam":std_data.slot_date,
            "time_of_exam":std_data.slot_time,
            "sections":datas,
            "total_questions":score_objs.totalquestions,
            "total_correct":score_objs.totalcorrectanswers,
            "total_incorrect":int(float(score_objs.totalquestionsattempted) - float(score_objs.totalcorrectanswers)),
            "total_not_attempted":int(float(score_objs.totalquestions) - float(score_objs.totalquestionsattempted)),

            "username": std_data.email,
            "user_id": std_data.id,
            "application_id": std_data.application_id,
            "student_name": f'''{std_data.first_name.upper()}" "{std_data.last_name.upper()}''',
            "slot_date": std_data.slot_date,
            "slot_time": std_data.slot_time,
            "photo": std_data.photo.url,
            "signature": std_data.signature.url,
            "barcode":"",
            "report_date": datetime.now(),
            # "test_link":"https://cocubes.in/gccschool-nfet",
            "test_link":"https://www.gccschool.com/myaccount",
            "bucket_static_logo":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/gcc-admit-card-logo.jpeg",
            # "bucket_static_signature":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/admit_card_signature.png"
            "bucket_static_signature":f"https://storage.googleapis.com/{static_selected_bucket}/static/images/gcc_admit_card_sign.png"
        }
        
        # Render template
        template = get_template("pdf/student_score_card.html")
        html = template.render(context)

        # xhtml2pdf needs ISO-8859-1
        html = html.encode("ISO-8859-1", "ignore").decode("ISO-8859-1")

        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_path = tmp.name
            pisa_status = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=tmp)

        if pisa_status.err:
            os.remove(pdf_path)
            return Response({"error": "PDF generation failed"}, status=500)
        try:
            # Upload to GCS
            username = re.sub(r"\s+", "_", std_data.email)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            gcs_file = f"media/admit_card/{username}_{std_data.id}.pdf"

            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)
            blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)






#########################################################################################


class StudentCreatePaymentView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        serializers = StudentCreatePaymentSerializer(data=request.data)
        if serializers.is_valid():
            obj = serializers.save()
            print("reqqq", request.user.email)
            lead = DossierData.objects.filter(email=request.user.email)
            print(lead)
            if lead:
                lead_obj = lead.last()
                Payments.objects.filter(id=obj.id).update(dossier_form=lead_obj, form_id=lead_obj.id, re_attempt_status=True)
            return Response({'message':'success',"status":200,'data':{}})
        else:
            return Response({'message':'failed',"status":400,'data':serializers.errors})
    


class PostExamResultView(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request):
        print("Result Request payload start...")
        print(request.data)
        print("Result Request payload end....")
        serializers = PostExamResultSerializer(data=request.data)
        if serializers.is_valid():
            datas = serializers.save()
            # print(datas)
            # print(type(datas))
            # print(datas.json_data)
            datas.json_data = request.data["competency"]
            datas.save()
            return Response({'message':'success',"status":200,'data':{}})
        else:
            return Response({'message':'failed',"status":400,'data':serializers.errors})
    

class PostRealExamResultView(APIView):
    # permission_classes = [IsAuthenticated]
    def post(self, request):
        print("Result Request payload start...")
        print(request.data)
        print("Result Request payload end....")
        serializers = PostRealExamResultSerializer(data=request.data)
        if serializers.is_valid():
            datas = serializers.save()
            # print(datas)
            # print(type(datas))
            # print(datas.json_data)
            datas.json_data = request.data["competency"]
            datas.save()
            return Response({'message':'success',"status":200,'data':{}})
        else:
            return Response({'message':'failed',"status":400,'data':serializers.errors})
    


### testing purpose

class AddWaiverValueProfileView(APIView):
    def post(self, request):
        dd = DossierData.objects.filter(document_status=2)
        print(len(dd))
        fee_waive_val = "Free of cost (FOC)"
        for i in dd:
            User.objects.filter(email=i.email).update(fee_waiver_category=fee_waive_val)
            DossierData.objects.filter(email=i.email, document_status=2).update(fee_waiver_category=fee_waive_val)


        # student = StudentProfile.objects.all()
        # for i in student:
            # ds = Payments.objects.filter(dossier_form__email=i.email, status="success").last()
            # if ds:
            #     i.fee_waiver_category = ds.fee_waiver_category
            #     i.save()
            #     print(i)
            # else:
            #     dd = DossierData.objects.filter(email=i.email).last()
            #     if dd:
            #         i.fee_waiver_category = dd.fee_waiver_category
            #         i.save()
            #         print(i)
            # ds = DossierData.objects.filter(email=i.email, document_status=2).last()
            ###########
            # if ds:
            #     fee_waive_val = "Free of cost (FOC)"
            #     i.fee_waiver_category = "Free of cost (FOC)"
            #     User.objects.filter(email=i.email).update(fee_waiver_category=fee_waive_val)
            #     DossierData.objects.filter(email=i.email, document_status=2).update(fee_waiver_category=fee_waive_val)
            #     i.save()
            #     print(i)

        return Response({'message':'success',"status":200,'data':{}})
    



## Add profile to meritto application form
from django.utils import timezone
import json
from datetime import datetime, timedelta, date

class AddProfileToMerittoView(APIView):
    def post(self, request):
        # datas = [
        #     "SANKT.SM@GMAIL.COM",
        #     "SAKSHISAIN07062004@GMAIL.COM",
        #     "BALAJILAXMI96@GMAIL.COM",
        #     "RAVIPANDEY6438@GMAIL.COM",
        #     "ANURAGSONI3032000@GMAIL.COM",
        #     "KALRASIDHI06@GMAIL.COM",
        #     "MAHEN3367@GMAIL.COM",
        #     "ONKARASWALE1998@GMAIL.COM",
        #     "AMANV95067@GMAIL.COM",
        #     "AAVEJGAVANDI5@GMAIL.COM",
        #     "CHARCHITBANSAL03@GMAIL.COM",
        #     "SAGNIKRAHA39@GMAIL.COM",
        #     "TAMANNAKHAN5405@GMAIL.COM",
        #     "AESONI99@GMAIL.COM",
        #     "AMANSOMANI00@GMAIL.COM",
        #     "NEERAJKASHYAPSINGH3@GMAIL.COM",
        #     "AYUSHCHANDEKAR8499@GMAIL.COM",
        #     "BKPRINCE1309@GMAIL.COM"
        # ]

        datas = [
            "vivekpandey804@gmail.com",
            "akshatsinha450@gmail.com",
            "pranjalpandey427@gmail.com",
            "kansalsaksham36@gmail.com",
            "sunakshiakshit@gmail.com",
            "aparnakumari12003@gmail.com",
            "dkritesh0904@gmail.com",
            "mehakkalsi2004@gmail.com",
            "agarwalrimjhim245@gmail.com",
            "goswamilokendra152@gmail.com",
            "prajapatidharmik455@gmail.com",
            "mdmahtabmalik8@gmail.com",
            "charangurrala01@gmail.com",
            "imohansingh967@gmail.com",
            "gautamsurtani@gmail.com",
            "krishnadwiv871@gmail.com",
            "gauravsharmajune5@gmail.com",
            "priyajindal1122@gmail.com",
            "chiragchandani2001@gmail.com",
            "pritykumari93346@gmail.com",
            "dksa085@gmail.com",
            "akilyacham97@gmail.com",
            "angelmaryj07@gmail.com",
            "ojus.ghosh3@gmail.com",
            "mfu212001@gmail.com",
            "hemanttm5@gmail.com",
            "spaavan4661@gmail.com",
            "radzshaw@gmail.com",
            "farooqui.m.atta@gmail.com",
            "faizalashfaquee@gmail.com",
            "yashmangal4u@gmail.com",
            "smarter05052002@gmail.com",
            "oms63613@gmail.com",
            "bcom.jitender@gmail.com",
            "pj131034@gmail.com",
            "singhikasak43@gmail.com",
            "kmittal2205@gmail.com",
            "iamasinghal.anish@gmail.com",
            "manmohini890@gmail.com",
            "navjotdran@gmail.com",
            "satya.choudhary7777@gmail.com",
            "khushpreetkaurmehe2904@gmail.com",
            "harshitsharma0413@gmail.com",
            "as8523055@gmail.com",
            "krishan.darc@gmail.com",
            "priyankababurao@protonmail.com",
            "sray11108@gmail.com",
            "sshaikhsimran26@gmail.com",
            "vg5311907@gmail.com",
            "khlnamdev@gmail.com",
            "gkkashyap0804@gmail.com",
            "vikashkumarmadheshiya612@gmail.com",
            "ghoshkrishnav@gmail.com",
            "sakshipalak2005@gmail.com",
            "varjatiayukti@gmail.com",
            "sainsenapati@gmail.com",
            "nrainkwar@gmail.com",
            "tanishka0416@gmail.com",
            "prachinoonwal3@gmail.com",
            "kanaksharma273@gmail.com",
            "saniyahussain906@gmail.com",
            "aryanjha333075@gmail.com",
            "siddhimahto111@gmail.com",
            "himanshubansalhb.13@gmail.com",
            "daiwikshah18@gmail.com",
            "gauravbarua570@gmail.com",
            "yogeshmishra88598@gmail.com",
            "mehuldonga745@gmail.com",
            "sachinkr051@gmail.com",
            "mehtaharshil207@gmail.com",
            "devahuja700@gmail.com",
            "ishikapawar4323@gmail.com",
            "ranapariyaparas01@gmail.com",
            "pramodin.yachuri@gmail.com",
            "patrosahil97@gmail.com",
            "upadhyaysomya210@gmail.com",
            "kumudkhushipandey@gmail.com",
            "chirag2005april@gmail.com",
            "kaushaltrivedi805@gmail.com",
            "pandeymayank1603@gmail.com",
            "ishagarg9145@gmail.com",
            "chadhadrishti643@gmail.com",
            "ishugautam849@gmail.com",
            "satish.singh010179@gmail.com",
            "vanisaini97@gmail.com",
            "shrutipoddar0808@gmail.com",
            "imranmughal06666@gmail.com",
            "samuelvincent104@gmail.com",
            "rashituli2005@gmail.com",
            "nishubishan21@gmail.com",
            "vidushi.sonker1995@gmail.com",
            "atulrajdnr@gmail.com",
            "sheeluyada.mba2025mb@rdias.ac.in",
            "amnplwl593@gmail.com",
            "tomarsakshi0007@gmail.com",
            "riyapundir33@gmail.com",
            "abhishekpradhan0501@gmail.com",
            "parulbagdi799@gmail.com",
            "sharma95305@gmail.com",
            "at6384582@gmail.com",
            "kanishkabansal898@gmail.com",
            "vanshbharti0001@gmail.com",
            "happytomar2003@gmail.com",
            "deepanshiy411@gmail.com",
            "janshiadhana@gmail.com",
            "singhrishika0717@gmail.com",
            "prachislnk23@gmail.com",
            "chestawadhwa729@gmail.com",
            "vishalsharma22309@gmail.com",
            "prateeksharma2401@gmail.com",
            "cprachi067@gmail.com",
            "sumantkunal4@gmail.com",
            "annujangid760@gmail.com",
            "sanjeetkr93112@gmail.com",
            "aadijain1967@gmail.com",
            "rashunarula18@gmail.com",
            "ys931137@gmail.com",
            "ctomar862@gmail.com",
            "sumitdangi317@gmail.com",
            "ajayshrivas0003@gmail.com",
            "ankitkumar7081723858@gmail.com",
            "raginiyadav96081@gmail.com",
            "aayushisuryavanshi9@gmail.com",
            "salonisjain1307@gmail.com",
            "rakshitjain974@gmail.com",
            "kauramanjot2723@gmail.com",
            "sehgalgirija@gmail.com",
            "amishadhiman946@gmail.com",
            "vaibhavkumar9909@gmail.com",
            "gurmannsingh1181@gmail.com",
            "nitikagrover.271@gmail.com",
            "ritikdhawan8@gmail.com",
            "abolirohakale431@gmail.com",
            "vermanaswag.999@gmail.com",
            "ny5429377@gmail.com",
            "k.r418765@gmail.com",
            "vanshitajarwal07@gmail.com",
            "ananthakrishnan.jayan.official@gmail.com",
            "shgarg700@gmail.com",
            "jyoti.shukla72165@gmail.com",
            "choprapayal590@gmail.com",
            "ss6305842@gmail.com",
            "sumanpreet3006@gmail.com",
            "shaikhadnan4460@gmail.com",
            "ipshitabajaj631@gmail.com",
            "singlanitish94@gmail.com",
            "kashish22rawat@gmail.com",
            "solankimaxx11@gmail.com",
            "prabhprabhjotkaur8@gmail.com",
            "richachaprana0@gmail.com",
            "ompatil8668@gmail.com",
            "anjaliraika2699@gmail.com",
            "shivamtyagi4157@gmail.com",
            "nagarsahil684@gmail.com",
            "souvikjpr2006@gmail.com",
            "anhalsuhani@gmail.com",
            "arjunverma2003.ca@gmail.com",
            "pallabdey2000@gmail.com",
            "rritu.5579@gmail.com",
            "me.vanshika.19@gmail.com",
            "sumitmh9@gmail.com",
            "nagardeep7017@gmail.com",
            "rahultiwary.du@gmail.com",
            "hsharma0679@gmail.com",
            "ishjotsinghbhatia@gmail.com",
            "anaspp489@gmail.com",
            "sandeepsnair1511@gmail.com",
            "kaursarbjit5437@gmail.com",
            "bipashadey55222@gmail.com",
            "aanchalmandloi104@gmail.com",
            "tomardeepika928@gmail.com",
            "ishitapanchal944@gmail.com",
            "vijayphule3115@gmail.com",
            "ajuusharma48@gmail.com",
            "shikhapnp2005@gmail.com",
            "priyanshitomar765@gmail.com",
            "shrutisharma1072@gmail.com",
            "khushikiran31@gmail.com",
            "sourabhbisht498@gmail.com",
            "charvimahajan2004@gmail.com",
            "mukhija.payal1123@gmail.com",
            "sumitsah2255@gmail.com",
            "kaursandeepkaur842@gmail.com",
            "satenderca@gmail.com",
            "kavyaagar.1998@gmail.com",
            "upendersharma87000@gmail.com",
            "shreesa02@gmail.com",
            "ahujadiya2675@gmail.com",
            "govindlangariya@gmail.com",
            "bhawandeepkaurg@gmail.com",
            "amankumar06@hotmail.com",
            "saveen1335@gmail.com",
            "katarampooja96@gmail.com",
            "ram024055@gmail.com",
            "priyanshum7617@gmail.com",
            "smit39813@gmail.com",
            "aniquashah51@gmail.com",
            "pranjalpariya10@gmail.com",
            "kavyatanwar18@gmail.com",
            "vedantb619@gmail.com",
            "aachalprajapati61@gmail.com",
            "revathisarai1214@gmail.com",
            "kajalsolanki661@gmail.com",
            "himanshichoudhary201@gmail.com",
            "chandangupta2901@gmail.com",
            "linileon30@gmail.com",
            "bansuri.choudharyy@gmail.com",
            "kajalrangeele@gmail.com",
            "kamaljitsinghratol@gmail.com",
            "shiva.mehta250@gmail.com",
            "rakshitkrrana@gmail.com",
            "vidushi8810@gmail.com",
            "dhritiman494@gmail.com",
            "kuwalamanvi@gmail.com",
            "305samirsvjc@gmail.com",
            "rafikali7035@gmail.com",
            "anushkadhillon18@gmail.com",
            "dsweety219@gmail.com",
            "singhleena725@gmail.com",
            "rajputsahil5521@gmail.com",
            "tushargupta0110@gmail.com",
            "infiniteshivam2004@gmail.com",
            "reedafatimakhan@gmail.com",
            "prachisharma1641@gmail.com",
            "abhi1109m@gmail.com",
            "dhruvguptaa2003@gmail.com",
            "alok11052004@gmail.com",
            "siddhiprasade09@gmail.com",
            "krishgoyal1310@gmail.com",
            "deepsimehta1234@gmail.com",
            "amanprazapat900@gmail.com",
            "tarmansharma2@gmail.com",
            "hardik14d@gmail.com",
            "raikriti632@gmail.com",
            "sneha.dwivedi052005@gmail.com",
            "anishbehera12@gmail.com",
            "sushilkr1493@gmail.com",
            "sidhaarthmr@gmail.com",
            "anureetbhinder23@gmail.com",
            "rajatbairolia1@gmail.com",
            "abdullahansari7838@gmail.com",
            "shreeguptatamanna@gmail.com",
            "officialid7310@gmail.com",
            "anchalg992@gmail.com",
            "vishalshah7270@gmail.com",
            "sanjanagill2001@gmail.com",
            "aditirohilla30@gmail.com",
            "nainagarg773@gmail.com",
            "abhimanyumishra1010@gmail.com",
            "preetix02004@gmail.com",
            "dkajudk5@gmail.com",
            "amansomani00@gmail.com",
            "ritikabazaz1@gmail.com",
            "rituu2899@gmail.com",
            "parisuryavanshi24@gmail.com",
            "arsh0306deep@gmail.com",
            "shaguns1628@gmail.com",
            "gharshita433@gmail.com",
            "dewangikaushish1200@gmail.com",
            "riddhimittal9136779415@gmail.com",
            "garima3112003@gmail.com",
            "kumkumsinngh@gmail.com",
            "info.aliasif113@gmail.com",
            "arjunborana321@gmail.com",
            "ritika.workbox@gmail.com",
            "roshaniverma482@gmail.com",
            "ishaan8775@gmail.com",
            "ashabhatt4959@gmail.com",
            "kesshavbareja@gmail.com",
            "pratishthagupta130@gmail.com",
            "vishwajeetkumarram80@gmail.com",
            "ibrahimansari9120@gmail.com",
            "anjanikashyap148@gmail.com",
            "lavishadixit@gmail.com",
            "samarthmavaghade@gmail.com",
            "devsharmasharma233@gmail.com",
            "jaygaveshnimavat@gmail.com",
            "atul.tevatia@kcglobed.com",
            "jashanpre257@gmail.com",
            "theayush2006@gmail.com",
            "valenciabharali05@gmail.com",
            "shraddhajais007@gmail.com",
            "dtanvi1027@gmail.com",
            "jaiveshpb13@gmail.com",
            "dhruvmittal8847@gmail.com",
            "ak3355438@gmail.com",
            "abhishekghosh574@gmail.com",
            "kumarkanhaiya977@gmail.com",
            "shaurya.vm09@gmail.com",
            "sushmajha0599@gmail.com",
            "khishi.250403@gmail.com",
            "lavanyachugh2005@gmail.com",
            "prathamarora665@gmail.com",
            "khateebr332@gmail.com",
            "devanshbharadwaj05@gmail.com",
            "rahulbhakat301@gmail.com",
            "skhesabuddin58@gmail.com",
            "aushmehta.lhf72@gmail.com",
            "abhinavanil1702@gmail.com",
            "kumarjhaabhishek2003@gmail.com",
            "kumarsu2601@gmail.com",
            "sakshibansal2103@gmail.com",
            "tarun1495@gmail.com",
            "nidhisingh92117@gmail.com",
            "singlalavita9@gmail.com",
            "adityarattan987@gmail.com",
            "sharmapreeti3244@gmail.com",
            "khadkeyadnesh29@gmail.com",
            "rg2608760@gmail.com",
            "riyatyagi.mba2025ea@rdias.ac.in",
            "jaman0677@gmail.com",
            "roushanshukla19@gmail.com",
            "akanshabhati066@gmail.com",
            "manikverma9840@gmail.com",
            "saievchavan@gmail.com",
            "sharathk.0801@gmail.com",
            "kherasoumya30@gmail.com",
            "shivanimandal191@gmail.com",
            "mitkarivrushali.june14@gmail.com",
            "shivanibasatiya@gmail.com",
            "sarikayadav0715@gmail.com",
            "prachigaur2215@gmail.com",
            "vickyyadav9692@gmail.com",
            "vinaysharma521411@gmail.com",
            "harshgoyal427289@gmail.com",
            "karimkhan3919@gmail.com",
            "prerna52970@gmail.com",
            "yshrajput1@gmail.com",
            "deevanshisardana.2021@gmail.com",
            "sakshiigupta006@gmail.com",
            "rk6609296@gmail.com",
            "shihabgazdhar@gmail.com",
            "samanpreetk245@gmail.com",
            "chadhagautam132@gmail.com",
            "tanishkpatil10b55@gmail.com",
            "yuvaveen@gmail.com",
            "yuviyuvarajyj@gmail.com",
            "padhiyarchirag1091@gmail.com",
            "raghavkuss2005@gmail.com",
            "rk5063196@gmail.com",
            "ajaykumarshrivas0003@gmail.com",
            "vivekjaiswal13032003@gmail.com",
            "adityasrivastav830@gmail.com",
            "sheetalsharma88261007@gmail.com",
            "nsingh84055@gmail.com",
            "adarshi2702@gmail.com",
            "tg83321@gmail.com",
            "rinkipal0102@gmail.com",
            "ks78656462@gmail.com",
            "varni.t.iwari0506@gmail.com",
            "sanjana885193@gmail.com",
            "manishnarayanamurthy@gmail.com",
            "iamkumkumrani@gmail.com",
            "zaheen8506@gmail.com",
            "thejas.v0204@gmail.com",
            "riyayadav0360@gmail.com",
            "baisakhidash20@gmail.com",
            "sumitpandey7415@gmail.com",
            "simrankaur302005@gmail.com",
            "anirudhbhardwaj2005@gmail.com",
            "shivamarora131324@gmail.com",
            "hemshikhathapa007@gmail.com",
            "sourabhvedi786@gmail.com",
            "premamayapadhi2@gmail.com",
            "pandeyvaishnawi0606@gmail.com",
            "samyakj573@gmail.com",
            "vidhigarg1411@gmail.com",
            "ar1176182@gmail.com",
            "sateeshpinninti3244@gmail.com",
            "tanania003@gmail.com",
            "anubhavsingh0664@gmail.com",
            "desaianushka28@gmail.com",
            "ganpathyiyer2899@gmail.com",
            "katariasimran406@gmail.com",
            "kmangalinaveen@gmail.com",
            "khanaatir868@gmail.com",
            "rathodnisha772@gmail.com",
            "pragyadalal4@gmail.com",
            "gulshanmehta1802@gmail.com",
            "talganrahul@gmail.com",
            "ayushshrivastava436@gmail.com",
            "deepdn1992@gmail.com",
            "lakhanbindal122@gmail.com",
            "samikshadu06@gmail.com",
            "sakshisain07062004@gmail.com",
            "aagarwal24680@gmail.com",
            "adityachandra71@gmail.com",
            "kalrasidhi06@gmail.com",
            "panchalchintu41@gmail.com",
            "khushishetty412@gmail.com",
            "ravipandey6438@gmail.com",
            "shabuddinmondal7667@gmail.com",
            "lovejeetkaur9881@gmail.com",
            "mashalsonia23@gmail.com",
            "ridhikalra346@gmail.com",
            "mahen3367@gmail.com",
            "connectasmit@gmail.com",
            "avniindoria2001@gmail.com",
            "devkaushik191@gmail.com",
            "aesoni99@gmail.com",
            "onkaraswale1998@gmail.com",
            "kizaz3625@gmail.com",
            "aavejgavandi5@gmail.com",
            "amanv95067@gmail.com",
            "charchitbansal03@gmail.com",
            "meghakardam@6gmail.com",
            "dhruvsingh9931@gmail.com",
            "simarpreetkaur0695@gmail.com",
            "prachi0564850@gmail.com",
            "dhruvbusiness006@gmail.com",
            "mallireddymaruthikumar@gmail.com",
            "verma.mansi9466@gmail.com",
            "apagrahari123@gmail.com",
            "piyushpopli2002@gmail.com",
            "tsingh9787@gmail.com",
            "essjayhora30@gmail.com",
            "bkprince1309@gmail.com",
            "ayushchandekar8499@gmail.com",
            "saniasol192@gmail.com",
            "angelroysms@gmail.com",
            "ky698625@gmail.com",
            "amishabhardwaj567@gmail.com",
            "souravsuper22@gmail.com",
            "sumitkumarjha3233@gmail.com",
            "nishkr29@gmail.com",
            "shivanishing973@gmail.com",
            "deepaksharma01205@gmail.com",
            "hanshika9696@gmail.com",
            "seemaguptaaug13@gmail.com"
        ]

        # datasp = []
        # for i in datas:
        #     datasp.append(i.lower())
        # print(datasp)
        # datasp = [emaill.lower() for emaill in datas]


        # datas = list(StudentRealExamResult.objects.all().values_list('student_profile__email', flat=True))

        # aapp_all = StudentProfile.objects.all().count()
        app_all = StudentProfile.objects.exclude(email__in=datas)
        # print("length....",len(app_all), "count",aapp_all, "alll result",len(datas))
        # for obj in app_all:
        #     if obj.slot_date:
        #         if obj.slot_date > datetime.now().date():
        #             status = "blank"
        #         if obj.slot_date == datetime.now().date():
        #             start_str, end_str = obj.slot_time.split(" - ")
        #             current_time = datetime.now().time().replace(microsecond=0)
        #             target_time = datetime.strptime(start_str, "%I:%M %p").time()
        #             dt1 = datetime.combine(date.today(), current_time)
        #             dt2 = datetime.combine(date.today(), target_time)

        #             print(dt1, dt2)
        #             if dt1>dt2:
        #                 current_time = datetime.now().time().replace(microsecond=0)
        #                 target_time = datetime.strptime(end_str, "%I:%M %p").time()
        #                 dt1 = datetime.combine(date.today(), current_time)
        #                 dt2 = datetime.combine(date.today(), target_time)
        #                 if dt1<dt2:
        #                     print(dt1, dt2)
        #                     status = "blank"
        #                 else:
        #                     std_exam  = ManageMasterKey.objects.filter(profile=obj.id)
        #                     if std_exam:
        #                         status = "Appeared"
        #                     else:
        #                         status = "Not Appeared"
        #             else:
        #                 std_exam  = ManageMasterKey.objects.filter(profile=obj.id)
        #                 if std_exam:
        #                     status = "Appeared"
        #                 else:
        #                     status = "blank"
        #         if obj.slot_date < datetime.now().date():
        #             std_exam  = ManageMasterKey.objects.filter(profile=obj.id)
        #             if std_exam:
        #                 status = "Appeared"
        #             else:
        #                 status = "Not Appeared"

        # return Response({"datas":""})
    

        # print("app data...",len(app_all))
        # for i in app_all:
        #     print(i.slot_date,'---', i.application_id, '----', i.email)
        # return Response({"start":"data"})
        print("total application", len(app_all))
        numm = 0
        success_datas = []
        failed_datas = []
        for query in app_all:
        #     print(query.slot_date)

        #     meritto_payload = {
        #             "form_id": 22144,
        #             # "email": "atul.tevatia@kcglobed.com",
        #             "email": query.user.email,
        #             "search_criteria": "email",
        #             "data": {
        #                 "field_351644" : "Appeared"
        #             }
        #         }
                
        #     print("meritto_payload...",meritto_payload)
        #     url = settings.MERITO_BASE_URL+"/application/v1/createOrUpdate"

        #     headers = {
        #             "Content-Type": "application/json",
        #             "secret-key": settings.MERITO_SECRETE_KEY,
        #             "access-key": settings.MERITO_ACCESS_KEY
        #         }

        #     try:
        #         response = requests.post(url, headers=headers, json=meritto_payload)
        #         print(response.status_code)
        #         print(response.text)
        #     except Exception as e:
        #         print("API Error:", str(e))
        # return Response({"datasss":""})
    
            if settings.MERITO_STATUS == "True":
                if int(query.gender) == 1:
                    mgender = "Male"
                elif int(query.gender) == 2:
                    mgender = "Female"
                else:
                    mgender = "Other"

                if int(query.tenth_medium) == 1:
                    mtmedium = "English"
                elif int(query.tenth_medium) == 2:
                    mtmedium = "Hindi"
                else:
                    mtmedium = "Other"

                if int(query.twelveth_medium) == 1:
                    mthmedium = "English"
                elif int(query.twelveth_medium) == 2:
                    mthmedium = "Hindi"
                else:
                    mthmedium = "Other"

                if int(query.medium_instruction) == 1:
                    minstrmedium = "English"
                elif int(query.medium_instruction) == 2:
                    minstrmedium = "Hindi"
                else:
                    minstrmedium = "Other"

                if query.higher_education_status == 1:
                    higher_status = "Yes"
                else:
                    higher_status = "No"

                if query.pg_status == 1:
                    pg_status = "Completed"
                else:
                    pg_status = "Pursuing"


                tenth_score_type = query.tenth_score_type if query.tenth_score_type == "Percentage" else "CGPA out of 10"
                twelveth_score_type = query.twelveth_score_type if query.twelveth_score_type == "Percentage" else "CGPA out of 10"

                if query.slot_time:
                    start_str, end_str = query.slot_time.split(" - ")
                    start_time = datetime.strptime(start_str, "%I:%M %p")
                    # Format to HH:mm:ss
                    start_formatted = start_time.strftime("%H:%M:%S %p")
                    start_formatted_one = start_time.strftime("%H:%M:%S")
                    # print(f'''{query.slot_date.strftime("%d/%m/%Y")} {start_formatted}''')

                if query.guardian_dropdown:
                    if int(query.guardian_dropdown) == 1:
                        gname = "Mother"
                    elif int(query.guardian_dropdown) == 2:
                        gname = "Father"
                    elif int(query.guardian_dropdown) == 3:
                        gname = "Other"
                    else:
                        gname = ""
                else:
                    gname = ""

                meritto_payload = {
                    "form_id": 22144,
                    "email": query.email,
                    "search_criteria":"email",
                    "data": {
                            "first_name":query.first_name,
                            "last_name":query.last_name,
                            "email":query.email,
                            "mobile_no":f"+91-{query.phone}",
                            "father_first_name":"",
                            "father_mobile_no":"",
                            "date_of_birth":query.date_of_birth.strftime("%d/%m/%Y"),
                            "gender":mgender,
                            "nationality":"Indian",
                            # "field_339552":query.state,
                            # "field_339553":query.city,
                            "field_337926":query.pincode,
                            # "field_340085":query.address,
                            # "field_340065":query.contact_name,
                            "field_340066":f"+91-{query.contact_phone}" if query.contact_phone else "",
                            "field_333993_1_1":query.tenth_passing_year,
                            "field_333993_1_2":tenth_score_type,
                            "field_333993_1_3":query.tenth_passing_percentage,
                            "field_333993_1_4":mtmedium,
                            "field_333994_1_1":query.twelveth_passing_year,
                            "field_333994_1_2":twelveth_score_type,
                            "field_333994_1_3":query.twelveth_passing_percentage,
                            "field_333994_1_4":mthmedium,
                            "field_340097_1_1":str(query.institution).replace("’",""),
                            "field_340097_1_2":query.ug_score_type,
                            "field_340097_1_3":format(float(query.pg_percentage), ".2f"),
                            "field_340097_1_4":format(float(query.pg_percentage), ".2f"),
                            "field_340069":pg_status,
                            "field_340077":higher_status,
                            "field_340079":query.higher_qualification_institution,
                            # "field_340078":query.higher_qualification,
                            "field_342113":query.user.application_id,
                            # "field_343097":"Complete",
                            "field_343098":"Complete",
                            "field_351358":query.guardian_name if query.guardian_name else "",
                            "field_351359":query.guardian_phone if query.guardian_phone else "",
                            "field_351368":query.guardian_email if query.guardian_email else "",
                            "field_351361":gname,
                            "field_351381":query.guardian_other_reason if query.guardian_other_reason else ""
                    }
                }

                if query.slot_time:
                    # meritto_payload["data"]["field_342102"] = start_formatted_one
                    meritto_payload["data"]["field_343386"] = f'''{query.slot_date.strftime("%d/%m/%Y")} {start_formatted}'''
                    meritto_payload["data"]["field_343097"] = "Complete"

                exp_payload = {"have_work_ex":"Fresher"}
                std_exp = StudentExperience.objects.filter(student_profile=query)
                if len(std_exp) > 0:
                    num = 1
                    exp_payload["have_work_ex"] = "Experienced"
                    for exp in std_exp:

                        key1 = f"field_334047_{num}_1"
                        value1 = exp.company_name
                        key2 = f"field_334047_{num}_2"
                        value2 = exp.position
                        key3 = f"field_334047_{num}_3"
                        value3 = exp.area
                        key4 = f"field_334047_{num}_4"
                        value4 = exp.start_date.strftime("%d/%m/%Y")
                        key5 = f"field_334047_{num}_5"
                        value5 = exp.end_date.strftime("%d/%m/%Y") if exp.end_date else exp.start_date.strftime("%d/%m/%Y")
                        key6 = f"field_334047_{num}_6"
                        value6 = ""

                        # print("values5...",value5)

                        exp_payload[key1] = value1
                        exp_payload[key2] = value2
                        exp_payload[key3] = value3
                        exp_payload[key4] = value4
                        exp_payload[key5] = value5
                        exp_payload[key6] = value6

                        print(exp_payload)
                        
                        num+=1


                print(exp_payload)
                meritto_payload["data"].update(exp_payload) 
                leads = list(DossierData.objects.filter(email=query.email).values_list('id'))
                payment_obj = Payments.objects.filter(dossier_form__in=leads, status="success")
                if payment_obj:
                    pay = payment_obj.first()
                    payment_payload = {
                        "field_342107":pay.razorpay_signature,
                        "field_342105":pay.razorpay_order_id,
                        "field_342106":pay.razorpay_payment_id,
                        "field_342108":int(pay.amount),
                        "field_342111":"INR",
                        "field_342110":pay.created_at.strftime("%d/%m/%Y %I:%M:%S %p"),
                        "field_342109":"success"
                    }
                    meritto_payload["data"].update(payment_payload)

                if query.slot_date:
                    if query.slot_date > datetime.now().date():
                        ap_status = "blank"
                    if query.slot_date == datetime.now().date():
                        start_str, end_str = query.slot_time.split(" - ")
                        current_time = datetime.now().time().replace(microsecond=0)
                        target_time = datetime.strptime(start_str, "%I:%M %p").time()
                        dt1 = datetime.combine(date.today(), current_time)
                        dt2 = datetime.combine(date.today(), target_time)

                        print(dt1, dt2)
                        if dt1>dt2:
                            current_time = datetime.now().time().replace(microsecond=0)
                            target_time = datetime.strptime(end_str, "%I:%M %p").time()
                            dt1 = datetime.combine(date.today(), current_time)
                            dt2 = datetime.combine(date.today(), target_time)
                            if dt1<dt2:
                                print(dt1, dt2)
                                ap_status = "blank"
                            else:
                                std_exam  = ManageMasterKey.objects.filter(profile=query.id)
                                if std_exam:
                                    ap_status = "Appeared"
                                else:
                                    ap_status = "Not Appeared"
                        else:
                            std_exam  = ManageMasterKey.objects.filter(profile=query.id)
                            if std_exam:
                                ap_status = "Appeared"
                            else:
                                ap_status = "blank"
                    if query.slot_date < datetime.now().date():
                        std_exam  = ManageMasterKey.objects.filter(profile=query.id)
                        if std_exam:
                            ap_status = "Appeared"
                        else:
                            ap_status = "Not Appeared"
                else:
                    ap_status = "blank"    
                # std_profile = StudentRealExamResult.objects.filter(student_profile=query)
                if ap_status == "blank":
                    meritto_payload["data"]["field_351644"] = ""
                else:
                    meritto_payload["data"]["field_351644"] = ap_status

                print("meritto_payload...",meritto_payload)
                
                url = settings.MERITO_BASE_URL+"/application/v1/createOrUpdate"

                headers = {
                        "Content-Type": "application/json",
                        "secret-key": settings.MERITO_SECRETE_KEY,
                        "access-key": settings.MERITO_ACCESS_KEY
                    }

                try:
                    response = requests.post(url, headers=headers, json=meritto_payload)
                    print(response.status_code)
                    print(response.text)
                    if int(response.status_code)==200:
                        success_datas.append(query.email)
                    else:
                        failed_datas.append(query.email)
                except Exception as e:
                    print("API Error:", str(e))
                    failed_datas.append(query.email)

                numm += 1
                print(numm)
        return Response({'message':'success',"status":200,'data':{"total":numm,"success_count":len(success_datas),"failed_count":len(failed_datas), "sucess":success_datas, "failed":failed_datas}})
    



##### Cocubes registartions:

from .services import CoCubesAssessmentService

class ScheduleAssessmentAPIView(APIView):

    def post(self, request):
        print("API Calling....!!")
        email = request.data.get("email")
        first_name = request.data.get("first_name")
        last_name = request.data.get("last_name", "")

        if not email or not first_name:
            return Response(
                {
                    "message": "email and first_name required",
                    "data":{}
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user_obj = StudentProfile.objects.filter(application_id=email)
            if user_obj:
                user_data = user_obj.first()

                # assigned_keys = ManageMasterKey.objects.filter(profile=user_data).values_list('key__key', flat=True)
                # available_pass_keys = ExamMasterKey.objects.filter(status=True).exclude(key__in=assigned_keys)
                # print("key dtaaa.,,,",available_pass_keys)
                status = False
                if not StudentRealExamResult.objects.filter(student_profile=user_data).exists():
                    day = datetime.now().day
                    available_pass_keys = ExamMasterKey.objects.filter(id=day, status=True)
                    if available_pass_keys:
                    #     available_pass_keys  = ExamMasterKey.objects.filter(status=True).first()
                    # else:
                        available_pass_keys = available_pass_keys.first()
                else:
                    assigned_keys = list(ManageMasterKey.objects.filter(profile=user_data).values_list('key__key', flat=True))
                    if assigned_keys:
                        available_pass_keys = ExamMasterKey.objects.filter(status=True).exclude(key__in=assigned_keys)
                        if available_pass_keys:
                            available_pass_keys = available_pass_keys.last()
                        else:
                            return Response({"status":404,"message":"Something went wrong","data":{}})
                    else:
                        available_pass_keys = ExamMasterKey.objects.filter(status=True).last()
                    status= True


                # print("data", type(available_pass_keys))
                # print("data2", len(available_pass_keys))
                # print("datas...",available_pass_keys)

                # ManageMasterKey.objects.create(profile=user_data, key=available_pass_keys, exam_url="qwertyuiop")
                # return Response({"msg":"success"})
            
                result = CoCubesAssessmentService.schedule_assessment(
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    # redirect_url="https://www.cocubes.com/",
                    redirect_url="https://www.gccschool.com/myaccount",
                    pass_keys = available_pass_keys.key
                )
                print(result)
                if result["erc"] == 0:
                    ManageMasterKey.objects.create(profile=user_data, key=available_pass_keys, exam_url=result["assessmentlink"], reattempt_status=status)
                    print("result..", result)
                    return Response({"status":200,"message":"Success","data":result})
                else:
                    print("start exam error",result)
                    return Response({"status":404,"message":result['err'],"data":{}})
            else:
                return Response({"status":404,"message":"Invalid User","data":{}})
        except Exception as e:
            return Response(
                {
                    "message": str(e),
                    "data":{}
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


