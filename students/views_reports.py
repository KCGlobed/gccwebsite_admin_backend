from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import *
from .serializers import *
from rest_framework import filters
from gcc_backend.pagination import CustomPageNumberPagination
from rest_framework.permissions import IsAuthenticated
from django.conf import settings
from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template
from google.cloud import storage
import os
import pandas as pd
import tempfile
import re
from datetime import datetime, timedelta
from django.utils.dateparse import parse_date
client = storage.Client(project=settings.GS_PROJECT_ID)
from django.db.models import F, FloatField, ExpressionWrapper
from django.db.models.functions import Cast

from openpyxl import load_workbook


class GetSessionReportPDFView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data_objs = CampusStudent.objects.all().order_by("-id")
        data_list = CampusStudentPDFSerializer(data_objs, many=True).data
        selected_bucket = settings.GS_BUCKET_NAME
        context = {
            "username": request.user.email,
            "user_id": request.user.id,
            "data_list": data_list,
            "report_date": datetime.now(),
            "bucket_static_logo":f"https://storage.googleapis.com/{selected_bucket}/static/images/gccschool.jpeg"
        }
        # return Response({"data":context})
        # Render template
        template = get_template("pdf/session_report.html")
        html = template.render(context)

        # xhtml2pdf needs ISO-8859-1
        html = html.encode("ISO-8859-1", "ignore").decode("ISO-8859-1")

        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_path = tmp.name
            pisa_status = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=tmp)

        if pisa_status.err:
            os.remove(pdf_path)
            return Response({"error": "PDF generation failed"}, status=500)

        try:
            # Upload to GCS
            username = re.sub(r"\s+", "_", request.user.email)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            gcs_file = f"media/pdf_reports/{username}_campusreport.pdf"

            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)
            blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)


class GetPaymentReportPDFView(APIView): 
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data_objs = Payments.objects.filter(source=SourceType.Website).order_by("-id")

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                data_objs = data_objs.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                data_objs = data_objs.filter(created_at__date__lte=end_date)

        data_list = ListPaymentPDFSerializer(data_objs, many=True).data
        selected_bucket = settings.GS_BUCKET_NAME
        context = {
            "username": request.user.email,
            "user_id": request.user.id,
            "data_list": data_list,
            "report_date": datetime.now(),
            "bucket_static_logo":f"https://storage.googleapis.com/{selected_bucket}/static/images/gccschool.jpeg"
        }
        # Render template
        template = get_template("pdf/payment_report.html")
        html = template.render(context)

        # xhtml2pdf needs ISO-8859-1
        html = html.encode("ISO-8859-1", "ignore").decode("ISO-8859-1")

        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_path = tmp.name
            pisa_status = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=tmp)

        if pisa_status.err:
            os.remove(pdf_path)
            return Response({"error": "PDF generation failed"}, status=500)
        try:
            # Upload to GCS
            username = re.sub(r"\s+", "_", request.user.email)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            gcs_file = f"media/pdf_reports/{username}_paymentreport.pdf"

            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)
            blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)





class GetSourcePaymentReportPDFView(APIView): 
    permission_classes = [IsAuthenticated]
    def get(self, request):
        
        source_type = request.GET.get("source")
        if source_type:
            data_objs = Payments.objects.filter(source=source_type).order_by('-id')
        else:
            data_objs = Payments.objects.filter(source=SourceType.Website).order_by('-id')
        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                data_objs = data_objs.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                data_objs = data_objs.filter(created_at__date__lte=end_date)

        data_list = ListPaymentPDFSerializer(data_objs, many=True).data
        selected_bucket = settings.GS_BUCKET_NAME
        context = {
            "username": request.user.email,
            "user_id": request.user.id,
            "data_list": data_list,
            "report_date": datetime.now(),
            "bucket_static_logo":f"https://storage.googleapis.com/{selected_bucket}/static/images/gccschool.jpeg"
        }
        # Render template
        template = get_template("pdf/payment_report.html")
        html = template.render(context)

        # xhtml2pdf needs ISO-8859-1
        html = html.encode("ISO-8859-1", "ignore").decode("ISO-8859-1")

        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_path = tmp.name
            pisa_status = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=tmp)

        if pisa_status.err:
            os.remove(pdf_path)
            return Response({"error": "PDF generation failed"}, status=500)
        try:
            # Upload to GCS
            username = re.sub(r"\s+", "_", request.user.email)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            gcs_file = f"media/pdf_reports/{username}_paymentreport.pdf"

            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)
            blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)





class GetCampusFacultyReportPDFView(APIView): 
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data_objs = CampusFaculty.objects.all().order_by("-id")

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                data_objs = data_objs.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                data_objs = data_objs.filter(created_at__date__lte=end_date)

        data_list = ListCampusFacultyPDFSerializer(data_objs, many=True).data
        selected_bucket = settings.GS_BUCKET_NAME
        context = {
            "username": request.user.email,
            "user_id": request.user.id,
            "data_list": data_list,
            "report_date": datetime.now(),
            "bucket_static_logo":f"https://storage.googleapis.com/{selected_bucket}/static/images/gccschool.jpeg"
        }
        # Render template
        template = get_template("pdf/campusfaculty_report.html")
        html = template.render(context)

        # xhtml2pdf needs ISO-8859-1
        html = html.encode("ISO-8859-1", "ignore").decode("ISO-8859-1")

        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_path = tmp.name
            pisa_status = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=tmp)

        if pisa_status.err:
            os.remove(pdf_path)
            return Response({"error": "PDF generation failed"}, status=500)

        try:
            # Upload to GCS
            username = re.sub(r"\s+", "_", request.user.email)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            gcs_file = f"media/pdf_reports/{username}_campusfaculty.pdf"

            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)
            blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)


### excel reports

class GetSessionReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        data_objs = CampusStudent.objects.all().order_by("-id")
        data_list = CampusStudentExcelSerializer(data_objs, many=True).data
        
        COLUMN_MAPPING = {
            "full_name": "Full Name",
            "email": "Email",
            "mobile": "Mobile",
            'city': 'City',
            'state': 'State',
            'student_reach': 'Reach Status',
            'address': 'Address',
            'college_name': 'College Name',
            'program_of_study': 'Program Of Study',
            'program_other': 'Other Program',
            'semester': 'Semester',
            'student_body_member': 'Student Body Member',
            'campus_ambassador_history': 'Campus Ambassador History',
            'inspiration': 'Inspiration'
            
            }
        # # Create temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            # Create DataFrame and save to the temporary file
            df = pd.DataFrame.from_dict(data_list)
            df.rename(columns=COLUMN_MAPPING, inplace=True)
            df.to_excel(pdf_path, header=True, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "campusreport"
            username = re.sub(r'\s+', '_', f"{request.user.first_name} {request.user.last_name}")
            gcs_folder_name = "media/excel_report"
            gcs_file_name = f"{gcs_folder_name}/{username}_{report_name}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)



class GetPaymentReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        # print("calling.. payment_report_excel")
        data_objs = Payments.objects.filter(source=SourceType.Website).order_by("-id")

        full_name = request.GET.get('full_name')
        if full_name:
            data_objs = data_objs.filter(dossier_form__full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            data_objs = data_objs.filter(dossier_form__email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            data_objs = data_objs.filter(dossier_form__phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            data_objs = data_objs.filter(dossier_form__state__icontains=state)

        city = request.GET.get('city')
        if city:
            data_objs = data_objs.filter(dossier_form__city__icontains=city)

        university = request.GET.get('university')
        if university:
            data_objs = data_objs.filter(dossier_form__university__icontains=university)
        fee_waiver_category = request.GET.get('fee_waiver_category')
        if fee_waiver_category:
            data_objs = data_objs.filter(fee_waiver_category__icontains=fee_waiver_category)

        status = request.GET.get('status')
        if status:
            data_objs = data_objs.filter(status__icontains=status)


        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                data_objs = data_objs.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                data_objs = data_objs.filter(created_at__date__lte=end_date)


        data_list = ListPaymentExcelReportSerializer(data_objs, many=True).data
        COLUMN_MAPPING = {
            "full_name":"Full Name",
            "email":"Email",
            "phone":"Phone",
            "city":"City",
            "state":"State",
            "university":"University",
            "ad_id":"Ad ID",
            "ad_source":"AD Source",
            "adset_id":"Adset ID",
            "campaign_id":"Campaign ID",
            "degree":"Degree",
            "degree_stage": "Degree Stage",
            "fbc_id": "FBC ID",
            "fbclid": "FBCLID",
            "utm_adname": "UTM Adname",
            "utm_campaign": "UTM Campaign",
            "utm_content": "UTM Content",
            "utm_medium": "UTM Medium",
            "utm_source": "UTM Source",
            "razorpay_order_id": "Order ID",
            "razorpay_payment_id": "Payment ID",
            "amount": "Amount",
            "status": "Payment Status",
            "created_at": "Payment Date",
            "fee_waiver_category":"Fee Waiver Category"
            }
        # # Create temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            #### Create DataFrame and save to the temporary file
            # df = pd.DataFrame.from_dict(data_list)
            # df.rename(columns=COLUMN_MAPPING, inplace=True)

            df = pd.DataFrame(data_list)

            # Reorder columns as per COLUMN_MAPPING keys
            df = df[list(COLUMN_MAPPING.keys())]

            # Rename columns for Excel headers
            df.rename(columns=COLUMN_MAPPING, inplace=True)

            df.to_excel(pdf_path, header=True, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "paymentreport"
            username = re.sub(r'\s+', '_', f"{request.user.first_name} {request.user.last_name}")
            gcs_folder_name = "media/excel_report"
            gcs_file_name = f"{gcs_folder_name}/{username}_{report_name}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)



class GetSourcePaymentReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        source_type = request.GET.get("source")
        if source_type:
            data_objs = Payments.objects.filter(source=source_type).order_by('-id')
        else:
            data_objs = Payments.objects.filter(source=SourceType.Website).order_by('-id')

        full_name = request.GET.get('full_name')
        if full_name:
            data_objs = data_objs.filter(dossier_form__full_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            data_objs = data_objs.filter(dossier_form__email__icontains=email)


        phone = request.GET.get('phone')
        if phone:
            data_objs = data_objs.filter(dossier_form__phone__icontains=phone)


        state = request.GET.get('state')
        if state:
            data_objs = data_objs.filter(dossier_form__state__icontains=state)

        city = request.GET.get('city')
        if city:
            data_objs = data_objs.filter(dossier_form__city__icontains=city)

        university = request.GET.get('university')
        if university:
            data_objs = data_objs.filter(dossier_form__university__icontains=university)
        fee_waiver_category = request.GET.get('fee_waiver_category')
        if fee_waiver_category:
            data_objs = data_objs.filter(fee_waiver_category__icontains=fee_waiver_category)

        status = request.GET.get('status')
        if status:
            data_objs = data_objs.filter(status__icontains=status)


        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                data_objs = data_objs.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                data_objs = data_objs.filter(created_at__date__lte=end_date)


        data_list = ListPaymentExcelReportSerializer(data_objs, many=True).data
        COLUMN_MAPPING = {
            "full_name":"Full Name",
            "email":"Email",
            "phone":"Phone",
            "city":"City",
            "state":"State",
            "university":"University",
            "ad_id":"Ad ID",
            "ad_source":"AD Source",
            "adset_id":"Adset ID",
            "campaign_id":"Campaign ID",
            "degree":"Degree",
            "degree_stage": "Degree Stage",
            "fbc_id": "FBC ID",
            "fbclid": "FBCLID",
            "utm_adname": "UTM Adname",
            "utm_campaign": "UTM Campaign",
            "utm_content": "UTM Content",
            "utm_medium": "UTM Medium",
            "utm_source": "UTM Source",
            "razorpay_payment_id": "Payment ID",
            "amount": "Amount",
            "status": "Payment Status",
            "created_at": "Payment Date",
            "fee_waiver_category":"Fee Waiver Category"
            }
        # # Create temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            #### Create DataFrame and save to the temporary file
            # df = pd.DataFrame.from_dict(data_list)
            # df.rename(columns=COLUMN_MAPPING, inplace=True)

            df = pd.DataFrame(data_list)

            # Reorder columns as per COLUMN_MAPPING keys
            df = df[list(COLUMN_MAPPING.keys())]

            # Rename columns for Excel headers
            df.rename(columns=COLUMN_MAPPING, inplace=True)

            df.to_excel(pdf_path, header=True, index=False)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "paymentreport"
            username = re.sub(r'\s+', '_', f"{request.user.first_name} {request.user.last_name}")
            gcs_folder_name = "media/excel_report"
            gcs_file_name = f"{gcs_folder_name}/{username}_{report_name}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)





class GetSessionFileUploadView(APIView):
    def post(self, request):
        file = request.FILES["file"]
        upload_for = request.data["upload_for"]  #images/files/videos
        bucket_flag = request.data["bucket"]

        # Upload to GCS
        if str(bucket_flag) == "1":
            gcs_file = f"media/{upload_for}/{file.name}"
            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)

            blob.upload_from_file(
            file,
            content_type=file.content_type
            )
            # blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "File uploaded successfully",
                "data": {
                    "file_name": file.name,
                    "gcs_path": gcs_file,
                    "download_url": url
                }
            })
        
        else:
            gcs_file = f"static/{upload_for}/{file.name}"
            bucket = client.bucket(settings.GS_BUCKET_NAME)
            blob = bucket.blob(gcs_file)
            blob.upload_from_file(
            file,
            content_type=file.content_type
            )
            # blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.public_url
            return Response({
                "message": "File uploaded successfully",
                "data": {
                    "file_name": file.name,
                    "gcs_path": gcs_file,
                    "download_url": url
                }
            })
        

### for student profile data
class GetStudentProfileReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        datas = StudentProfile.objects.all().order_by("-id")

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        # slot_date range filter
        start_slot_date = request.GET.get('start_slot_date')
        end_slot_date = request.GET.get('end_slot_date')
        if start_slot_date:
            datas = datas.filter(slot_date__gte=start_slot_date)

        if end_slot_date:
            datas = datas.filter(slot_date__lte=end_slot_date)


        fee_waiver_category = request.GET.get('fee_waiver_category')
        if fee_waiver_category:
            datas = datas.filter(fee_waiver_category=fee_waiver_category)

        first_name = request.GET.get('first_name')
        if first_name:
            datas = datas.filter(first_name__icontains=first_name)
        last_name = request.GET.get('last_name')
        if last_name:
            datas = datas.filter(last_name__icontains=last_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)
        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)
        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)
        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)
        is_result = request.GET.get('is_result')

        # print(is_result)
        if is_result:
            # print(is_result)
            if is_result == "50":

                datas = datas.filter(
                        id__in=StudentRealExamResult.objects.annotate(
                            total_score_float=Cast('totalscore', FloatField()),
                            total_questions_float=Cast('totalquestions', FloatField()),
                        ).annotate(
                            percentage=ExpressionWrapper(
                                (F('total_score_float') * 100.0) / F('total_questions_float'),
                                output_field=FloatField()
                            )
                        ).filter(
                            percentage__gt=50
                        ).values_list(
                            "student_profile_id",
                            flat=True
                        )
                    )
            else:
                # print("datass")
                datas = datas.filter(
                    id__in=StudentRealExamResult.objects.values_list(
                        "student_profile_id",
                        flat=True
                    )
                )

        data_list = ListStudentProfileExcelReportSerializer(datas, many=True).data
        # print("datas...",data_list)
        COLUMN_MAPPING = {
            "first_name":"First Name",
            "last_name":"Last Name",
            "email":"Email",
            "phone":"Phone Number",
            "contact_name":"Contact Name",
            "contact_phone": "Contact Phone Number",
            "date_of_birth": "Date Of Birth",
            "gender": "Gender",
            "nationality": "Nationality",
            "pincode": "PinCode",
            "city": "City",
            "state": "State",
            "address": "Full Address",
            "tenth_passing_year": "Year Of Pass(10th)",
            "tenth_passing_percentage": "Percentage/CGPA(10th)",
            "tenth_score_type": "Score Type(10th)",
            "tenth_medium": "Medium(10th)",
            "twelveth_passing_year": "Year Of Pass(12th)",
            "twelveth_passing_percentage": "Percentage/CGPA(12th)",
            "twelveth_score_type": "Score Type(12th)",
            "twelveth_medium": "Medium(12th)",
            "medium_instruction": "Medium(UG)",
            "pg_status":"Study(UG)",
            "ug_score_type": "Percentage/CGPA(UG)",
            "pg_percentage": "Score(UG)",
            "institution": "Institution(UG)",
            "higher_education_status": "Higher Qualification Status",
            "higher_qualification": "Higher Qualification",
            "higher_qualification_institution": "Higher Qualification Institution",
            "employement_status":"Experience",
            "slot_date": "Slot Date",
            "slot_time": "Slot Time",
            "student_result":"Student Result",
            "application_id": "Application ID",
            "created_at": "Create Date",
            "fee_waiver_category": "Fee Waiver Category",
            "referral_code": "Refferal Code",
            "referred_code": "Reffered Code",
            "guardian_name": "Guardian Name",
            "guardian_phone": "Guardian Phone",
            "guardian_email": "Guardian Email",
            "guardian_dropdown": "Guardian Relationship",
            "resume":"Resume"
            }
        # print(data_list)
        # return Response({"message":"success", "data":data_list})
        # # Create temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            #### Create DataFrame and save to the temporary file
            # df = pd.DataFrame.from_dict(data_list)
            # df.rename(columns=COLUMN_MAPPING, inplace=True)

            df = pd.DataFrame(data_list)

            # Reorder columns as per COLUMN_MAPPING keys
            df = df[list(COLUMN_MAPPING.keys())]

            # Rename columns for Excel headers
            df.rename(columns=COLUMN_MAPPING, inplace=True)

            # df.to_excel(pdf_path, header=True, index=False)
            df.to_excel(pdf_path, header=True, index=False)

            # Make Resume column clickable
            wb = load_workbook(pdf_path)
            ws = wb.active

            resume_column = None
            
            # Find Resume column
            for col_num, cell in enumerate(ws[1], start=1):
                if cell.value == "Resume":
                    resume_column = col_num
                    break

            # Add hyperlink
            if resume_column:
                for row_num in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=resume_column)

                    if cell.value:
                        resume_url = str(cell.value).strip()

                        if resume_url.startswith(("http://", "https://")):
                            cell.hyperlink = resume_url
                            cell.value = "View Resume"  # Optional
                            cell.style = "Hyperlink"

            wb.save(pdf_path)
            
            # result_column = None
            
            # # Find Resume column
            # for col_num, cell in enumerate(ws[1], start=1):
            #     if cell.value == "Score Card Result":
            #         result_column = col_num
            #         # print("result...", result_column)
            #         break

            # # Add hyperlink
            # if result_column:
            #     for row_num in range(2, ws.max_row + 1):
            #         cell = ws.cell(row=row_num, column=result_column)

            #         if cell.value:
            #             result_url = str(cell.value).strip()

            #             if result_url.startswith(("http://", "https://")):
            #                 cell.hyperlink = result_url
            #                 cell.value = "View Result"  # Optional
            #                 cell.style = "Hyperlink"

            # wb.save(pdf_path)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "student_profile_report"
            username = re.sub(r'\s+', '_', f"{request.user.first_name} {request.user.last_name}")
            gcs_folder_name = "media/excel_report"
            gcs_file_name = f"{gcs_folder_name}/{username}_{report_name}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)




class GetStudentProfileReportPDFView(APIView): 
    permission_classes = [IsAuthenticated]
    def get(self, request):
        datas = StudentProfile.objects.all().order_by("-id")

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)


        # slot_date range filter
        start_slot_date = request.GET.get('start_slot_date')
        end_slot_date = request.GET.get('end_slot_date')
        if start_slot_date:
            datas = datas.filter(slot_date__gte=start_slot_date)

        if end_slot_date:
            datas = datas.filter(slot_date__lte=end_slot_date)

        
        fee_waiver_category = request.GET.get('fee_waiver_category')
        if fee_waiver_category:
            datas = datas.filter(fee_waiver_category=fee_waiver_category)

        first_name = request.GET.get('first_name')
        if first_name:
            datas = datas.filter(first_name__icontains=first_name)
        last_name = request.GET.get('last_name')
        if last_name:
            datas = datas.filter(last_name__icontains=last_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(email__icontains=email)
        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(phone__icontains=phone)
        state = request.GET.get('state')
        if state:
            datas = datas.filter(state__icontains=state)
        city = request.GET.get('city')
        if city:
            datas = datas.filter(city__icontains=city)



        data_list = ListStudentProfileExcelReportSerializer(datas, many=True).data
        selected_bucket = settings.GS_BUCKET_NAME
        context = {
            "username": request.user.email,
            "user_id": request.user.id,
            "data_list": data_list,
            "report_date": datetime.now(),
            "bucket_static_logo":f"https://storage.googleapis.com/{selected_bucket}/static/images/gccschool.jpeg"
        }
        # Render template
        template = get_template("pdf/student_profile_report.html")
        html = template.render(context)

        # xhtml2pdf needs ISO-8859-1
        html = html.encode("ISO-8859-1", "ignore").decode("ISO-8859-1")

        # Create temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_path = tmp.name
            pisa_status = pisa.CreatePDF(BytesIO(html.encode("ISO-8859-1")), dest=tmp)

        if pisa_status.err:
            os.remove(pdf_path)
            return Response({"error": "PDF generation failed"}, status=500)
        try:
            # Upload to GCS
            username = re.sub(r"\s+", "_", request.user.email)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            gcs_file = f"media/pdf_reports/{username}_profilereport.pdf"

            bucket = client.bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file)
            blob.upload_from_filename(pdf_path, content_type="application/pdf")
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)


######################## Interview ######################


### for student profile Interview data
class GetStudentProfileInterviewReportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        datas = ManageStudentInterview.objects.all().order_by("-id")

        # Date range filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            start_date = parse_date(start_date)
            if start_date:
                datas = datas.filter(created_at__date__gte=start_date)

        if end_date:
            end_date = parse_date(end_date)
            if end_date:
                datas = datas.filter(created_at__date__lte=end_date)

        # first_name = request.GET.get('first_name')
        # if first_name:
        #     datas = datas.filter(profile__first_name__icontains=first_name)

        # last_name = request.GET.get('last_name')
        # if last_name:
        #     datas = datas.filter(profile__last_name__icontains=last_name)
        full_name = request.GET.get('full_name')
        if full_name:
            datas = datas.filter(profile__user__first_name__icontains=full_name)

        email = request.GET.get('email')
        if email:
            datas = datas.filter(profile__email__icontains=email)

        phone = request.GET.get('phone')
        if phone:
            datas = datas.filter(profile__phone__icontains=phone)

        source = request.GET.get('source')
        if source:
            dossier_datas = list(DossierData.objects.filter(source=source).values_list('email', flat=True))
            datas = datas.filter(profile__email__in=dossier_datas)
            
        data_list = StudentInterviewReportSerializer(datas, many=True).data
        # print("datas...",data_list)
        COLUMN_MAPPING = {
            "first_name":"First Name",
            "last_name":"Last Name",
            "email":"Email",
            "phone":"Phone Number",
            "application_id":"Student ID",
            "interview_date": "Interview Schedule Date",
            "attempt_status": "Attendence",
            "company_name": "Company Name"
            }
        # print(data_list)
        # return Response({"message":"success", "data":data_list})
        # # Create temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            pdf_path = temp_file.name
            
            #### Create DataFrame and save to the temporary file
            # df = pd.DataFrame.from_dict(data_list)
            # df.rename(columns=COLUMN_MAPPING, inplace=True)

            df = pd.DataFrame(data_list)

            # Reorder columns as per COLUMN_MAPPING keys
            df = df[list(COLUMN_MAPPING.keys())]

            # Rename columns for Excel headers
            df.rename(columns=COLUMN_MAPPING, inplace=True)

            # df.to_excel(pdf_path, header=True, index=False)
            df.to_excel(pdf_path, header=True, index=False)

            # Make Resume column clickable
            wb = load_workbook(pdf_path)
            # ws = wb.active

            # resume_column = None
            
            # # Find Resume column
            # for col_num, cell in enumerate(ws[1], start=1):
            #     if cell.value == "Company Name":
            #         resume_column = col_num
            #         break

            # # Add hyperlink
            # if resume_column:
            #     for row_num in range(2, ws.max_row + 1):
            #         cell = ws.cell(row=row_num, column=resume_column)

            #         if cell.value:
            #             resume_url = cell.value
            #             cell.value = cell.value["name"]
            #             # if resume_url.startswith(("http://", "https://")):
            #             #     cell.hyperlink = resume_url
            #             #     # cell.value = "View Resume"  # Optional
            #             #     cell.value = "View Resume"
            #             #     # cell.style = "Hyperlink"

            # wb.save(pdf_path)
        
        # After the 'with' block, the file is closed but not deleted
        try:
            # GCS file naming logic
            timestamp = datetime.now().strftime("%d_%m_%y_%H_%M")
            report_name = "student_interview_report"
            username = re.sub(r'\s+', '_', f"{request.user.first_name} {request.user.last_name}")
            gcs_folder_name = "media/excel_report"
            gcs_file_name = f"{gcs_folder_name}/{username}_{report_name}.xlsx"

            # Upload the temporary file to GCS
            bucket = client.get_bucket(settings.GS_BUCKET_NAME_2)
            blob = bucket.blob(gcs_file_name)
            blob.upload_from_filename(pdf_path)
            # ---------- Generate signed URL ----------
            url = blob.generate_signed_url(
                version="v4",
                expiration=timedelta(minutes=settings.SIGNED_URL_EXPIRY),
                method="GET"
            )
            return Response({
                "message": "Success",
                "data": {
                    "report_url": url
                }
            })

        finally:
            os.remove(pdf_path)

